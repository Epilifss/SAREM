# Bibliotecas
import tkinter as tk
import ttkbootstrap as tb
import threading
import time
import pyodbc
import hashlib
import sys
import os
import subprocess
import tempfile
import datetime
import matplotlib.pyplot as plt
import pandas as pd
import configparser
import requests
from tkinter import messagebox, filedialog
from ttkbootstrap import Style
from ttkbootstrap.constants import *
from moviepy import VideoFileClip
from PIL import Image, ImageTk
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.ticker import FuncFormatter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_LEFT
from reportlab.platypus import Paragraph
from reportlab.lib.units import cm
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from textwrap import wrap

# Componentes
from components import *

# Modulos
from modulos.corporativo import CorporativoModule
from modulos.varejo import VarejoModule
from modulos.exportacao import ExportacaoModule

# Banco de dados
from database import create_connection_Protheus
from database import create_connection

# Estilos
from theme import temas, estilizar_treeview, ajustar_largura_colunas

# Funções

def center_window_screen(win, width=350, height=100):
    win.update_idletasks()
    screen_width = win.winfo_screenwidth()
    screen_height = win.winfo_screenheight()
    x = (screen_width // 2) - (width // 2)
    y = (screen_height // 2) - (height // 2)
    win.geometry(f"{width}x{height}+{x}+{y}")

def resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

VERSAO_ATUAL = "1.0.3"
URL_VERSAO = "https://dl.dropboxusercontent.com/scl/fi/dh936k1ph0m2eq5low9gn/versao.txt?rlkey=fnyxw89yee7ai571ue3znbaiv&st=xd9f26ij&dl=0"
URL_LINK_TXT = "https://dl.dropboxusercontent.com/scl/fi/9hu4p9wo6ydqiit7zfp1b/linknovaversao.txt?rlkey=yubwve92t8qpr2bxd13lsvqij&st=vs0ddwfv&dl=0"

def get_download_url():
    try:
        resposta = requests.get(URL_LINK_TXT, timeout=5)
        if resposta.status_code == 200:
            return resposta.text.strip()
    except Exception as e:
        print("Erro ao obter link do instalador:", e)
    return None

def verificar_e_att(parent):
    try:
        resposta = requests.get(URL_VERSAO, timeout=5)
        if resposta.status_code == 200:
            versao_remota = resposta.text.strip()
            if versao_remota > VERSAO_ATUAL:
                resp = messagebox.askyesno(
                    "SAREM - Atualização disponível",
                    f"Uma nova versão está disponível.\nDeseja atualizar agora?"
                )
                if not resp:
                    return True

                url_download = get_download_url()
                if not url_download:
                    messagebox.showerror("Erro", "Não foi possível obter o link do instalador.")
                    return True

                # Janela de progresso
                progress_win = tk.Toplevel(parent)
                progress_win.title("Baixando atualização")
                progress_win.iconbitmap(resource_path('SAREM.ico'))
                progress_win.geometry("350x100")
                progress_win.transient(parent)
                progress_win.grab_set()
                tb.Label(progress_win, text="Baixando nova versão...").pack(pady=10)
                progress = tb.Progressbar(progress_win, orient="horizontal", length=300, mode="determinate")
                progress.pack(pady=10)
                progress["value"] = 0

                center_window_screen(progress_win, 350, 100)

                temp_dir = tempfile.gettempdir()
                caminho_instalador = os.path.join(temp_dir, "atualizacao.exe")
                r = requests.get(url_download, stream=True)
                total = int(r.headers.get('content-length', 0))
                chunk_size = 8192
                baixado = 0
                f = open(caminho_instalador, 'wb')
                chunks = r.iter_content(chunk_size=chunk_size)

                def baixar_pedaco():
                    nonlocal baixado
                    try:
                        chunk = next(chunks)
                        if chunk:
                            f.write(chunk)
                            baixado += len(chunk)
                            if total:
                                valor = (baixado / total) * 100
                                progress["value"] = valor
                            progress_win.update_idletasks()
                        progress_win.after(1, baixar_pedaco)
                    except StopIteration:
                        f.close()
                        progress_win.destroy()
                        if os.path.exists(caminho_instalador):
                            messagebox.showinfo("Atualização", "Download concluído! O instalador será aberto.")
                            print(f"Tentando abrir o instalador: {caminho_instalador}")
                            try:
                                subprocess.Popen(['explorer', caminho_instalador])
                            except Exception as e:
                                messagebox.showerror("Erro", f"Não foi possível abrir o instalador:\n{e}")
                            parent.quit()
                        else:
                            messagebox.showerror("Erro", "O instalador não foi baixado corretamente.")

                baixar_pedaco()
                progress_win.transient()
                progress_win.grab_set()
                progress_win.wait_window()
                return False

        return True
    except requests.RequestException as e:
        print(f"Erro ao verificar atualização: {e}")
        return True

monitorar_bo_event = threading.Event()

def monitorar_bo_embarcadas():
    def tarefa():
        while not monitorar_bo_event.is_set():
            try:
                horario = datetime.datetime.now()
                hora = (f"as {horario.hour}:{horario.minute}")

                print("Monitorando BOs embarcadas " + hora)
                conn = create_connection()
                conn_Protheus = create_connection_Protheus()
                if conn is None or conn_Protheus is None:
                    print("Erro de conexão com o banco.")
                    time.sleep(60)
                    continue

                cursor = conn.cursor()
                cursor.execute("SELECT bo_number, op FROM bo_records WHERE status <> 'Embarcado'")
                bos = cursor.fetchall()

                for bo_number, op in bos:
                    cursor_mik = conn_Protheus.cursor()
                    cursor_mik.execute("""
                        SELECT COUNT(*) FROM SC6010
                        WHERE C6_NUM = ? AND C6_BLQ = '' AND D_E_L_E_T_ <> '*'
                    """, str(op))
                    total_itens = cursor_mik.fetchone()[0]

                    cursor_mik.execute("""
                        SELECT COUNT(*) FROM SC6010
                        WHERE C6_NUM = ? AND C6_BLQ = '' AND D_E_L_E_T_ <> '*' AND C6_NOTA <> ''
                    """, str(op))
                    itens_com_nota = cursor_mik.fetchone()[0]
                    cursor_mik.close()

                    if total_itens > 0 and total_itens == itens_com_nota:
                        cursor.execute("UPDATE bo_records SET status = 'Embarcado' WHERE bo_number = ?", bo_number)
                        conn.commit()
                        messagebox.showinfo("BO Liberada para Embarque", f"A {bo_number} foi liberada para embarque.")
                        print(f"BO {bo_number} marcada como Embarcada.")

                        try:
                            from modulos.corporativo import CorporativoModule
                            if hasattr(CorporativoModule, "instance") and CorporativoModule.instance is not None:
                                CorporativoModule.instance.atualizar_bos()
                        except Exception:
                            pass
                        try:
                            from modulos.varejo import VarejoModule
                            if hasattr(VarejoModule, "instance") and VarejoModule.instance is not None:
                                VarejoModule.instance.atualizar_bos()
                        except Exception:
                            pass
                        try:
                            from modulos.exportacao import ExportacaoModule
                            if hasattr(ExportacaoModule, "instance") and ExportacaoModule.instance is not None:
                                ExportacaoModule.instance.atualizar_bos()
                        except Exception:
                            pass

                cursor.close()
                conn.close()
                conn_Protheus.close()
            except Exception as e:
                print(f"Erro ao monitorar BOs embarcadas: {e}")
            if monitorar_bo_event.wait(60):
                print("Monitoramento de BOs embarcadas interrompido.")
                break

    monitorar_bo_event.clear()
    threading.Thread(target=tarefa, daemon=True).start()

def center_window(self):
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        self.root.geometry(f"{width}x{height}+{x}+{y}")

def center_window_child(child, parent):
    child.update_idletasks()
    width = child.winfo_width()
    height = child.winfo_height()
    parent_x = parent.winfo_rootx()
    parent_y = parent.winfo_rooty()
    parent_width = parent.winfo_width()
    parent_height = parent.winfo_height()

    x = parent_x + (parent_width // 2) - (width // 2)
    y = parent_y + (parent_height // 2) - (height // 2)
    child.geometry(f"{width}x{height}+{x}+{y}")

class TreeViewHelper:
    def ajustar_colunas(self):
        for col in self.tree["columns"]:
            valores = [self.tree.set(item, col) for item in self.tree.get_children()]
        if not valores:
            largura = 50  # valor padrão mínimo
        else:
            largura = max(len(str(valor)) for valor in valores) * 10
        self.tree.column(col, width=largura) 

def get_config_path():
    appdata = os.environ.get('APPDATA') or os.path.expanduser('~')
    config_dir = os.path.join(appdata, 'SAREM')
    os.makedirs(config_dir, exist_ok=True)
    return os.path.join(config_dir, 'config.ini')

CONFIG_PATH = get_config_path()

class ConfigEditor:
    def __init__(self, parent_login=None):
        self.parent_login = parent_login
        self.root = tk.Toplevel(parent_login)
        self.root.title("Configuração do Banco de Dados")
        self.root.iconbitmap(resource_path('SAREM.ico'))
        self.root.minsize(400, 300)
        self.root.resizable(False, False)
        self.root.transient(parent_login)
        self.root.grab_set()

        self.root.protocol("WM_DELETE_WINDOW", self.cancelar)

        self.config = configparser.ConfigParser()
        self.config.read(CONFIG_PATH)
        with open(CONFIG_PATH, "w") as configfile:
            self.config.write(configfile)

        # Campos para o banco SAREM
        campos_db = [
            ("server", "Servidor"),
            ("database", "Banco de Dados"),
            ("user", "Usuário"),
            ("password", "Senha")
        ]
        # Campos para o banco Protheus
        campos_prot = [
            ("server", "Servidor"),
            ("database", "Banco"),
            ("user", "Usuário"),
            ("password", "Senha")
        ]
        self.entries_db = {}
        self.entries_mik = {}

        frame = tb.Frame(self.root, padding=20)
        frame.pack(expand=True, fill=tk.BOTH, anchor="center")

        # SAREM
        tb.Label(frame, text="Banco de Dados SAREM", font=("Arial", 10, "bold")).pack(pady=(0, 10))
        for campo, label in campos_db:
            row = tb.Frame(frame)
            row.pack(fill="x", pady=3)
            tb.Label(row, text=label + ":", width=15, anchor="w").pack(side="left")
            entry = tb.Entry(row, width=30, show="*" if campo == "password" else "")
            entry.pack(side="left", padx=5)
            self.entries_db[campo] = entry

        # Protheus
        tb.Label(frame, text="Banco Protheus", font=("Arial", 10, "bold")).pack(pady=(20, 10))
        for campo, label in campos_prot:
            row = tb.Frame(frame)
            row.pack(fill="x", pady=3)
            tb.Label(row, text=label + ":", width=15, anchor="w").pack(side="left")
            entry = tb.Entry(row, width=30, show="*" if campo == "password" else "")
            entry.pack(side="left", padx=5)
            self.entries_mik[campo] = entry

        # Carrega valores atuais
        if "database" in self.config:
            for campo in self.entries_db:
                valor = self.config["database"].get(campo, "")
                self.entries_db[campo].insert(0, valor)
        if "Protheus" in self.config:
            for campo in self.entries_mik:
                valor = self.config["Protheus"].get(campo, "")
                self.entries_mik[campo].insert(0, valor)

        # Botões
        btn_frame = tb.Frame(frame)
        btn_frame.pack(pady=15)
        tb.Button(btn_frame, text="Salvar", command=self.salvar, style="custom.TButton", width=15).pack(side="left", padx=10)
        tb.Button(btn_frame, text="Cancelar", command=self.cancelar, style="custom.TButton", width=15).pack(side="left", padx=10)

        center_window(self)
        self.root.after(100, lambda: self.entries_db["server"].focus_set())

    def cancelar(self):
        self.root.destroy()
        try:
            if self.parent_login:
                self.parent_login.deiconify()
        except Exception:
            pass

    def salvar(self):
        if "database" not in self.config:
            self.config["database"] = {}
        if "Protheus" not in self.config:
            self.config["Protheus"] = {}
        for campo, entry in self.entries_db.items():
            self.config["database"][campo] = entry.get()
        for campo, entry in self.entries_mik.items():
            self.config["Protheus"][campo] = entry.get()
        try:
            with open(CONFIG_PATH, "w") as configfile:
                self.config.write(configfile)
            messagebox.showinfo("Sucesso", "Configuração salva com sucesso!")
            self.root.destroy()
            try:
                if self.parent_login:
                    self.parent_login.deiconify()
            except Exception:
                pass
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao salvar: {e}")

class DialogSenhaAdmin:
    def __init__(self, parent):
        self.result = None
        self.root = tk.Toplevel(parent)
        self.root.title("Senha de administrador")
        self.root.iconbitmap(resource_path('SAREM.ico'))
        self.root.transient(parent)
        self.root.focus_set()
        self.root.resizable(False, False)

        frame = tb.Frame(self.root, padding=20)
        frame.pack(expand=True, fill="both")

        tb.Label(frame, text="Digite a senha de administrador para acessar as configurações:").pack(pady=(0, 10))
        self.entry = tb.Entry(frame, show="*", width=25)
        self.entry.pack(pady=5)
        self.entry.focus_set()

        btn_frame = tb.Frame(frame)
        btn_frame.pack(pady=10)
        tb.Button(btn_frame, text="OK", width=10, command=self.ok, style="custom.TButton").pack(side="left", padx=5)
        tb.Button(btn_frame, text="Cancelar", width=10, command=self.cancel, style="custom.TButton").pack(side="left", padx=5)

        self.root.bind('<Return>', lambda event: self.ok())
        self.root.bind('<Escape>', lambda event: self.cancel())

        center_window(self)
        self.root.wait_window()

    def ok(self):
        self.result = self.entry.get()
        self.root.destroy()

    def cancel(self):
        self.result = None
        self.root.destroy()

class LoginWindow:
    def __init__(self, parent):
        self.parent = parent
        self.root = tk.Toplevel(parent)
        self.root.title("SAREM")
        self.root.iconbitmap(resource_path('SAREM.ico'))
        # self.root.geometry("350x300")
        self.root.geometry("1000x600")
        self.root.state('zoomed')
        self.root.focus_set()

        self.root.protocol("WM_DELETE_WINDOW", parent.destroy)

        self.frame = tb.Frame(self.root, padding=10)
        self.frame.pack(expand=True, fill="both")

        # Frame Campos de entrada
        campos_frame = tb.Frame(self.frame)
        campos_frame.pack(pady=200, expand=True, anchor="center")
        
        rodape_frame = tb.Frame(self.frame)
        rodape_frame.pack(pady=0, expand=True, anchor="se")
        tb.Label(rodape_frame, text=f"Versão {VERSAO_ATUAL}").pack(side="right", padx=5)

        logo_img = Image.open(resource_path('SAREM PNG.png'))
        logo_img = logo_img.resize((250, 80), Image.LANCZOS)
        self.logo_photo = ImageTk.PhotoImage(logo_img)
        tb.Label(campos_frame, image=self.logo_photo).pack(pady=(0, 10))


        # Usuário
        user_text = tb.Frame(campos_frame)
        user_text.pack(fill="both", pady=1, expand=True, side="top")
        tb.Label(user_text, text="Usuário:", width=10, anchor="w").pack(side="left", padx=0)
        # Entry para o usuário
        user_entry = tb.Frame(campos_frame)
        user_entry.pack(fill="both", pady=1, expand=True)
        self.username = tb.Entry(user_entry, width=40)
        self.username.pack(side="left", padx=0)

        # Senha
        pass_text = tb.Frame(campos_frame)
        pass_text.pack(fill="both", pady=1, expand=True, side="top")
        tb.Label(pass_text, text="Senha:", width=10, anchor="w").pack(side="left")
        # Entry para a senha
        pass_entry = tb.Frame(campos_frame)
        pass_entry.pack(fill="both", pady=1)
        self.password = tb.Entry(pass_entry, show="*", width=40)
        self.password.pack(side="left", padx=0)

        # Botões
        btn_frame = tb.Frame(campos_frame)
        btn_frame.pack(pady=5, expand=True, anchor="center")

        self.login_btn = tb.Button(btn_frame, text="Login", width=17, command=self.login)
        self.login_btn.pack(side="left", padx=1)

        self.config_btn = tb.Button(btn_frame, text="Configurar Banco", width=17, command=self.abrir_config)
        self.config_btn.pack(side="right", padx=1)

        center_window(self)
        self.username.focus_set()
        self.root.bind('<Return>', self.login)

    def abrir_config(self):
        dialog = DialogSenhaAdmin(self.root)
        senha = dialog.result
        correta = "4dmin123@4"
        if senha == "":
            messagebox.showerror("Erro", "Insira uma senha para continuar")
        elif senha == None:
            return
        elif senha != correta:
            messagebox.showerror("Acesso negado", "Senha incorreta!")
        elif senha == correta:
            ConfigEditor(self.root)
    
    def admin_login(self):
        self.username.delete(0, tk.END)
        self.password.delete(0, tk.END)

    def login(self, event=None):
        config = configparser.ConfigParser()
        config.read(CONFIG_PATH)
        campos_obrigatorios = ["server", "database", "user", "password"]
        if "database" not in config or any(not config["database"].get(campo, "").strip() for campo in campos_obrigatorios):
            messagebox.showwarning(
                "Configuração necessária",
                "Preencha as configurações do banco de dados antes de fazer login."
            )
            self.abrir_config()
            return

        username = self.username.get()
        password = self.password.get()
        hashed_pw = hashlib.sha256(password.encode()).hexdigest()
        

        conn = create_connection()
        if conn is None:
            messagebox.showerror("Erro", "Falha ao conectar ao banco de dados")
            return

        try:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT * FROM users WHERE username COLLATE Latin1_General_BIN=? AND password_hash COLLATE Latin1_General_BIN=?", (username, hashed_pw))
            user = cursor.fetchone()
            

            if user:
                printUser = (f"Usuário {user[1]} logou no ")

                horario = datetime.datetime.now()
                hora = (f"as {horario.hour}:{horario.minute}")

                if user[4]:  # is_admin
                    self.root.destroy()
                    print(printUser + "painel de administração " + hora)
                    AdminPanel(self.parent)
                else:
                    module = user[3]
                    if (module) == '0':
                        self.root.destroy()
                        print(printUser + "módulo Corporativo " + hora)
                        monitorar_bo_embarcadas()
                        CorporativoModule(user, self.parent)
                    elif (module) == '1':
                        self.root.destroy()
                        print(printUser + "módulo Varejo " + hora)
                        monitorar_bo_embarcadas()
                        VarejoModule(user, self.parent)
                    elif (module) == '2':
                        self.root.destroy()
                        print(printUser + "módulo Exportação " + hora)
                        monitorar_bo_embarcadas()
                        ExportacaoModule(user, self.parent)
                    else:
                        messagebox.showerror(
                            "Erro", "Módulo do usuário não encontrado. Contate o administrador.")
            else:
                messagebox.showerror("Erro", "Credenciais inválidas")
        except pyodbc.Error as e:
            messagebox.showerror(
                "Erro", f"Erro ao consultar o banco de dados: {e}")
        finally:
            if conn:
                cursor.close()
                conn.close()


class AdminPanel:
    def __init__(self, parent):
        self.root = tk.Toplevel(parent)
        self.root.title("Painel de Administração")
        self.root.iconbitmap(resource_path('SAREM.ico'))

        self.root.protocol("WM_DELETE_WINDOW", parent.destroy)

        self.root.state('zoomed')

        # Abas
        self.notebook = tb.Notebook(self.root)

        # Aba de Usuários
        self.user_frame = tb.Frame(self.notebook)
        self.notebook.add(self.user_frame, text="Usuários")

        header = tb.Frame(self.user_frame)
        header.pack(fill=tk.X)

        tb.Button(header, text="Novo Usuário",
                   command=self.novo_usuario).pack(side=tk.LEFT)
        tb.Button(header, text="Excluir Usuário",
                   command=self.excluir_usuario).pack(side=tk.LEFT)
        tb.Button(header, text="Editar Usuário",
                   command=self.editar_usuario).pack(side=tk.LEFT)
        tb.Button(header, text="Logoff",
                   command=self.logoff).pack(side=tk.RIGHT)
        
        # Lista de usuários
        self.tree = tb.Treeview(self.user_frame, columns=(
            "ID", "Usuário", "Módulo", "Admin"), show="headings")
        self.tree.heading("ID", text="ID")
        self.tree.heading("Usuário", text="Usuário")
        self.tree.heading("Módulo", text="Módulo")
        self.tree.heading("Admin", text="Admin")
        self.tree.column("ID", width=0, stretch=False)
        self.tree.pack(expand=True, fill=tk.BOTH)

        self.tree.bind("<<TreeviewSelect>>", self.atualizar_selecao)

        # Aba de Database
        self.db_frame = tb.Frame(self.notebook)
        self.notebook.add(self.db_frame, text="Banco de Dados")
        self.criar_formulario_db()

        self.notebook.pack(expand=True, fill=tk.BOTH)
        self.carregar_usuarios()

    def carregar_usuarios(self):
        conn = create_connection()
        if conn is None:
            return

        try:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT id, username, IIF(module='0','Corporativo',IIF(module='1','Varejo',IIF(module='2', 'Exportação', ''))), IIF(is_admin='True', 'Sim', 'Não') FROM users")
            rows = cursor.fetchall()

            for item in self.tree.get_children():
                self.tree.delete(item)

            for row in rows:
                valores_formatados = [str(item) for item in row]
                self.tree.insert("", tk.END, values=valores_formatados)
        except pyodbc.Error as e:
            messagebox.showerror("Erro", f"Erro ao carregar usuários: {e}")
        finally:
            if conn:
                cursor.close()
                conn.close()

    def atualizar_selecao(self, event=None):
        selected_items = self.tree.selection()
        if selected_items:
            item_id = selected_items[0]
            valores = self.tree.item(item_id, "values")
            self.usuario_selecionado = valores[0]
            self.nome_usuario = valores[1]
        else:
            self.usuario_selecionado = None

    def novo_usuario(self):
        NovoUsuarioWindow(self.root, self)

    def excluir_usuario(self):
        if not hasattr(self, 'usuario_selecionado') or not self.usuario_selecionado:
            messagebox.showerror("Erro", "Nenhum usuário selecionado")
            return

        resposta = messagebox.askyesno(
            "Confirmar", f"Tem certeza que deseja excluir o usuário {self.nome_usuario}?")
        if not resposta:
            return

        conn = create_connection()
        if conn is None:
            messagebox.showerror("Erro", "Falha ao conectar ao banco de dados")
            return

        try:
            cursor = conn.cursor()

            cursor.execute("SELECT * FROM users WHERE id=?",
                           (self.usuario_selecionado,))
            usuario_existente = cursor.fetchone()

            if not usuario_existente:
                messagebox.showerror(
                    "Erro", f"O usuário '{self.usuario_selecionado}' não foi encontrado no banco de dados.")
                return

            usuario_formatado = self.usuario_selecionado.strip()

            print(f"Excluindo usuário: {usuario_formatado}")

            cursor.execute("DELETE FROM users WHERE id=?",
                           (self.usuario_selecionado,))
            conn.commit()

            if cursor.rowcount == 0:
                messagebox.showerror(
                    "Erro", "Nenhum registro foi excluído. Verifique se o usuário existe.")
            else:
                messagebox.showinfo("Sucesso", "Usuário excluído com sucesso!")

            self.carregar_usuarios()
            self.usuario_selecionado = None

        except pyodbc.Error as e:
            messagebox.showerror("Erro", f"Erro ao excluir usuário: {e}")

        finally:
            if conn:
                cursor.close()
                conn.close()

    def editar_usuario(self):
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showwarning(
                "Atenção", "Selecione um usuário para editar.")
            return

        user_data = self.tree.item(selected_items[0], "values")
        EditarUsuarioWindow(self.root, self, user_data)

    def criar_formulario_db(self):
        CONFIG_PATH = get_config_path()
        self.db_config = configparser.ConfigParser()
        self.db_config.read(CONFIG_PATH)

        campos_db = [
            ("server", "Servidor"),
            ("database", "Banco de Dados"),
            ("user", "Usuário"),
            ("password", "Senha")
        ]
        campos_prot = [
            ("server", "Servidor"),
            ("database", "Banco"),
            ("user", "Usuário"),
            ("password", "Senha")
        ]
        self.db_entries = {}
        self.mik_entries = {}

        frame = tb.Frame(self.db_frame, padding=20)
        frame.pack(expand=True, fill=tk.BOTH)

        tb.Label(frame, text="Banco de Dados SAREM", font=("Arial", 10, "bold")).grid(row=0, column=0, columnspan=2, pady=(0, 10))
        for i, (campo, label) in enumerate(campos_db):
            tb.Label(frame, text=label + ":").grid(row=i+1, column=0, sticky=tk.W, pady=5)
            entry = tb.Entry(frame, width=30, show="*" if campo == "password" else "")
            entry.grid(row=i+1, column=1, pady=5)
            self.db_entries[campo] = entry

        tb.Label(frame, text="Banco Protheus", font=("Arial", 10, "bold")).grid(row=len(campos_db)+1, column=0, columnspan=2, pady=(20, 10))
        for i, (campo, label) in enumerate(campos_prot):
            tb.Label(frame, text=label + ":").grid(row=len(campos_db)+2+i, column=0, sticky=tk.W, pady=5)
            entry = tb.Entry(frame, width=30, show="*" if campo == "password" else "")
            entry.grid(row=len(campos_db)+2+i, column=1, pady=5)
            self.mik_entries[campo] = entry

        if "database" in self.db_config:
            for campo in self.db_entries:
                valor = self.db_config["database"].get(campo, "")
                self.db_entries[campo].insert(0, valor)
        if "Protheus" in self.db_config:
            for campo in self.mik_entries:
                valor = self.db_config["Protheus"].get(campo, "")
                self.mik_entries[campo].insert(0, valor)

        row_btn = len(campos_db) + len(campos_prot) + 2
        tb.Button(frame, text="Salvar", command=self.salvar_config_db).grid(row=row_btn, column=0, pady=15, sticky="ew")
        tb.Button(frame, text="Cancelar", command=self.cancelar_config_db).grid(row=row_btn, column=1, pady=15, sticky="ew")

    def salvar_config_db(self):
        CONFIG_PATH = get_config_path()
        if "database" not in self.db_config:
            self.db_config["database"] = {}
        if "Protheus" not in self.db_config:
            self.db_config["Protheus"] = {}
        for campo, entry in self.db_entries.items():
            self.db_config["database"][campo] = entry.get()
        for campo, entry in self.mik_entries.items():
            self.db_config["Protheus"][campo] = entry.get()
        try:
            with open(CONFIG_PATH, "w") as configfile:
                self.db_config.write(configfile)
            messagebox.showinfo("Sucesso", "Configuração salva com sucesso!")
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao salvar: {e}")

    def cancelar_config_db(self):
        for campo, entry in self.db_entries.items():
            entry.delete(0, tk.END)
            valor = self.db_config["database"].get(campo, "")
            entry.insert(0, valor)
        for campo, entry in self.mik_entries.items():
            entry.delete(0, tk.END)
            valor = self.db_config["Protheus"].get(campo, "")
            entry.insert(0, valor)

    def logoff(self, parent=None):
        global monitorar_bo_event
        monitorar_bo_event.set()
        if parent is None:
            parent = self.root
        self.root.destroy()
        login_win = LoginWindow(parent.master if hasattr(parent, 'master') else parent)
        login_win.root.after(100, lambda: (login_win.root.focus_force(), login_win.username.focus_set()))


class NovoUsuarioWindow:
    def __init__(self, parent, admin_panel):
        self.root = tk.Toplevel(parent)
        self.root.title("Novo Usuário")
        self.root.iconbitmap(resource_path('SAREM.ico'))
        self.root.grab_set()
        self.root.focus_set()

        self.admin_panel = admin_panel

        campos = ["Usuário", "Senha", "Módulo", "Admin"]
        self.entries = {}

        content = tb.Frame(self.root, padding=20)
        content.grid(row=0, column=0, sticky="nsew")
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

        for i, campo in enumerate(campos):
            tb.Label(content, text=f"{campo}:").grid(
            row=i, column=0, sticky=tk.W, padx=10, pady=8
            )
            if campo == "Senha":
                entry = tb.Entry(content, show="*")
            elif campo == "Admin":
                entry = tb.Combobox(content, values=["Sim", "Não"], state="readonly")
            elif campo == "Módulo":
                entry = tb.Combobox(content, values=[
                "Corporativo", "Varejo", "Exportação"], state="readonly")
            else:
                entry = tb.Entry(content)

            entry.grid(row=i, column=1, padx=10, pady=8)
            self.entries[campo] = entry

        tb.Button(content, text="Salvar", command=self.salvar).grid(
            row=len(campos), column=0, columnspan=2, pady=(15, 0)
        )

        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f"{width}x{height}+{x}+{y}")

    def salvar(self):
        conn = create_connection()
        if conn is None:
            return

        try:
            cursor = conn.cursor()
            username = self.entries["Usuário"].get()
            password = self.entries["Senha"].get()
            hashed_pw = hashlib.sha256(password.encode()).hexdigest()
            module = self.entries["Módulo"].get()
            is_admin = 1 if self.entries["Admin"].get() == "Sim" else 0
            module = 0 if self.entries["Módulo"].get(
            ) == "Corporativo" else 1 if self.entries["Módulo"].get() == "Varejo" else 2

            cursor.execute('''
                INSERT INTO users (username, password_hash, module, is_admin)
                VALUES (?, ?, ?, ?)
            ''', (username, hashed_pw, module, is_admin))

            conn.commit()
            messagebox.showinfo("Sucesso", "Usuário cadastrado com sucesso!")
            self.root.destroy()

            self.admin_panel.carregar_usuarios()
        except pyodbc.Error as e:
            messagebox.showerror("Erro", f"Erro ao salvar usuário: {e}")
        finally:
            if conn:
                cursor.close()
                conn.close()


class EditarUsuarioWindow:
    def __init__(self, parent, admin_panel, user_data):
        self.root = tk.Toplevel(parent)
        self.root.title("Editar Usuário")
        self.root.iconbitmap(resource_path('SAREM.ico'))
        self.root.grab_set()
        self.root.focus_set()
        self.admin_panel = admin_panel
        self.user_data = user_data

        campos = ["Usuário", "Senha", "Módulo", "Admin"]
        self.entries = {}

        content = tb.Frame(self.root, padding=20)
        content.grid(row=0, column=0, sticky="nsew")
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

        for i, campo in enumerate(campos):
            tb.Label(content, text=f"{campo}:").grid(
            row=i, column=0, sticky=tk.W, padx=10, pady=8
            )

            if campo == "Admin":
                entry = tb.Combobox(content, values=["Sim", "Não"], state="readonly")
                entry.set("Sim" if self.user_data[3] == "Sim" else "Não")
            elif campo == "Senha":
                entry = tb.Entry(content, show="*")
            elif campo == "Módulo":
                entry = tb.Combobox(content, values=[
                    "Corporativo", "Varejo", "Exportação"], state="readonly")
                entry.set(self.user_data[2])
            else:
                entry = tb.Entry(content)
            if campo == "Usuário":
                entry.insert(0, self.user_data[1])
            entry.grid(row=i, column=1, padx=10, pady=8)
            self.entries[campo] = entry

        tb.Button(content, text="Salvar", command=self.salvar).grid(
            row=len(campos), column=0, columnspan=2, pady=(15, 0)
        )

        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f"{width}x{height}+{x}+{y}")

    def salvar(self):
        conn = create_connection()
        if conn is None:
            return

        try:
            cursor = conn.cursor()
            username = self.entries["Usuário"].get()
            new_password = self.entries["Senha"].get()
            module = self.entries["Módulo"].get()
            is_admin = 1 if self.entries["Admin"].get() == "Sim" else 0
            module = 0 if self.entries["Módulo"].get(
            ) == "Corporativo" else 1 if self.entries["Módulo"].get() == "Varejo" else 2

            if new_password:
                hashed_pw = hashlib.sha256(new_password.encode()).hexdigest()
                cursor.execute('''
                    UPDATE users
                    SET username = ?, password_hash = ?, module = ?, is_admin = ?
                    WHERE id = ?
                ''', (username, hashed_pw, module, is_admin, self.user_data[0]))  # ID está no índice 0
            else:
                cursor.execute('''
                    UPDATE users
                    SET username = ?, module = ?, is_admin = ?
                    WHERE id = ?
                ''', (username, module, is_admin, self.user_data[0]))  # ID está no índice 0

            conn.commit()
            messagebox.showinfo("Sucesso", "Usuário editado com sucesso!")
            self.root.destroy()

            self.admin_panel.carregar_usuarios()
        except pyodbc.Error as e:
            messagebox.showerror("Erro", f"Erro ao editar usuário: {e}")
        finally:
            if conn:
                cursor.close()
                conn.close()


if __name__ == "__main__":
    LoginWindow()


class Embarcados(TreeViewHelper):
    def __init__(self, parent, user, caller_id=None):
        self.parent = parent
        self.user = user
        self.root = tk.Toplevel(parent)
        self.root.title("Embarcados")
        self.root.iconbitmap(resource_path('SAREM.ico'))
        self.root.geometry("1000x600")

        self.root.transient(parent)
        center_window_child(self.root, parent)

        self.ultimo_modulo = caller_id
        print(self.identificar_chamador())

        # Cabeçalho
        header = tb.Frame(self.root)
        header.pack(fill=tk.X)

        tb.Button(header, text="Fechar",
                   command=lambda: self.root.destroy()).pack(side=tk.RIGHT)

        # Barra de pesquisa
        self.search_bar = SearchBar(
            self.root, self.pesquisar_bo, self.clear_search)

        # Lista de BOs
        self.tree = tb.Treeview(self.root, columns=(
            "BO", "OP", "Status", "tipo_ocorrencia", "dt_embarque"), show="headings")
        self.tree.heading("BO", text="BO")
        self.tree.heading("OP", text="OP")
        self.tree.heading("Status", text="Status")
        self.tree.heading("tipo_ocorrencia", text="Tipo de Ocorrência")
        self.tree.heading("dt_embarque", text="Data de Embarque")

        self.tree.pack(expand=True, fill=tk.BOTH)

        self.tree.bind("<Double-1>", lambda event: exibir_detalhes_embarcado(self.root,
                       self.tree, caller_id="Embarcados"))

        self.carregar_bos()

    def identificar_chamador(self):
        if self.ultimo_modulo is None:
            raise ValueError(
                "É necessário informar o identificador do módulo que chamou a função")
        return f"Tela de embarcados chamada pelo módulo: {self.ultimo_modulo}"

    def carregar_bos(self):
        conn = create_connection()
        if conn is None:
            return

        try:
            cursor = conn.cursor()
            query = """
                SELECT bo_number, op, status, tipo_ocorrencia, dt_embarque FROM bo_records WHERE status LIKE 'Embarcado' AND modulo LIKE ?
            """
            cursor.execute(query, self.ultimo_modulo)
            rows = cursor.fetchall()

            # Remove todos os itens atuais da Treeview
            for item in self.tree.get_children():
                self.tree.delete(item)

            # Insere os dados "limpos" na Treeview
            for row in rows:
                # Converte cada valor para string (se não for None) e aplica strip para remover caracteres indesejados.
                bo = str(row[0]).strip("(' ,)") if row[0] is not None else ""
                op = str(row[1]).strip("(' ,)") if row[1] is not None else ""
                status = str(row[2]).strip(
                    "(' ,)") if row[2] is not None else ""
                tipo_ocorrencia = str(row[3]).strip(
                    "(' ,)") if row[3] is not None else ""
                motivo = str(row[4]).strip(
                    "(' ,)") if row[4] is not None else ""

                self.tree.insert("", tk.END, values=(
                    bo, op, status, tipo_ocorrencia, motivo))

            self.ajustar_colunas()
        except pyodbc.Error as e:
            messagebox.showerror("Erro", f"Erro ao carregar BOs: {e}")
        finally:
            if conn:
                cursor.close()
                conn.close()

    def pesquisar_bo(self, event=None):
        termo = self.search_bar.search_entry.get()
        if not termo:
            return

        conn = create_connection()
        if conn is None:
            return

        try:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT bo_number, op, status, tipo_ocorrencia, motivo FROM bo_records WHERE bo_number LIKE ?", (f"%{termo}%",))
            rows = cursor.fetchall()

            self.tree.delete(*self.tree.get_children())
            for row in rows:
                # Converte cada valor para string (se não for None) e aplica strip para remover caracteres indesejados.
                bo = str(row[0]).strip("(' ,)") if row[0] is not None else ""
                op = str(row[1]).strip("(' ,)") if row[1] is not None else ""
                status = str(row[2]).strip(
                    "(' ,)") if row[2] is not None else ""
                tipo_ocorrencia = str(row[3]).strip(
                    "(' ,)") if row[3] is not None else ""
                dt_embarque = str(row[4]).strip(
                    "(' ,)") if row[4] is not None else ""

                self.tree.insert("", tk.END, values=(
                    bo, op, status, tipo_ocorrencia, dt_embarque))
                
            self.ajustar_colunas()
        except pyodbc.Error as e:
            messagebox.showerror("Erro", f"Erro ao pesquisar BOs: {e}")
        finally:
            if conn:
                cursor.close()
                conn.close()

    def clear_search(self):
        self.search_bar.search_entry.delete(0, tk.END)
        self.search_bar.update_buttons()
        self.carregar_bos()

class Estatisticas:
    def __init__(self, parent, caller_id=None):
        self.parent = parent
        self.root = tk.Toplevel(parent)
        self.root.title("Estatísticas")
        self.root.iconbitmap(resource_path('SAREM.ico'))
        self.root.geometry("800x500")
        self.root.state('zoomed')

        self.root.transient(parent)
        center_window_child(self.root, parent)

        self.ultimo_modulo = caller_id
        print(self.identificar_chamador())
        
        self.anos = self.obter_anos()
        self.setores = self.obter_setores("Todos")
        self.contagens = self.obter_contagens_por_setor()
        self.ano_atual = "Todos"

        # Cabeçalho
        header = tb.Frame(self.root)
        header.pack(fill=tk.X, padx=10, pady=10)

        tb.Label(header, text="Ano: ").pack(side=tk.LEFT)
        self.listaAnos = tb.Combobox(header, values=self.anos, state="readonly")
        self.listaAnos.pack(side=tk.LEFT)
        self.listaAnos.set(self.ano_atual)
        self.listaAnos.bind("<<ComboboxSelected>>", self.atualizar_grafico)
        
        tb.Button(header, text="Fechar",
                   command=lambda: self.root.destroy()).pack(side=tk.RIGHT)
        
        # Gráfico
        self.frame = tb.Frame(self.root, padding="3 3 12 12")
        self.frame.pack(fill=tk.BOTH, expand=True)
        self.fig = Figure(figsize=(8, 6), dpi=100)
        self.ax = self.fig.add_subplot(111)
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        self.atualizar_grafico()

        
    def identificar_chamador(self):
        if self.ultimo_modulo is None:
            raise ValueError(
                "É necessário informar o identificador do módulo que chamou a função")
        return f"Tela de estatísticas chamada pelo módulo: {self.ultimo_modulo}"

    def obter_anos(self):
        conn = create_connection()
        cursor = conn.cursor()

        cursor.execute(f"SELECT DISTINCT YEAR(emissao_totvs) AS ano FROM bo_records WHERE emissao_totvs IS NOT NULL AND modulo like ? ORDER BY ano DESC", (self.ultimo_modulo,))
        anos = ["Todos"] + [row[0] for row in cursor.fetchall()]

        cursor.close()
        conn.close()

        return anos
        
    def obter_setores(self, ano):
        conn = create_connection()
        cursor = conn.cursor()

        if ano == "Todos":
            query = ("SELECT COALESCE(setor_responsavel, 'Não especificado') FROM bo_records WHERE modulo like ?")
            cursor.execute(query, self.ultimo_modulo)
        else:
            query = ("SELECT COALESCE(setor_responsavel, 'Não especificado') FROM bo_records WHERE YEAR(emissao_totvs) = ? AND modulo like ? ORDER BY COALESCE(setor_responsavel, 'Não especificado')")
            cursor.execute(query, (ano, self.ultimo_modulo))
        setores = [row[0] for row in cursor.fetchall()]

        cursor.close()
        conn.close()

        return setores
        
    def obter_contagens_por_setor(self):
        conn = create_connection()
        cursor = conn.cursor()

        cursor.execute(f"SELECT COALESCE(setor_responsavel, 'Não especificado'), COUNT(*) FROM bo_records GROUP BY setor_responsavel")
        contagens = {row[0]: row[1] for row in cursor.fetchall()}
        cursor.close()
        conn.close()

        return contagens

    def atualizar_grafico(self, event=None):
        self.ano_atual = self.listaAnos.get()
        self.setores = self.obter_setores(self.ano_atual)  # Obter setores com base no filtro selecionado
        self.contagens = self.obter_contagens_por_setor_por_ano(self.ano_atual)  # Obter contagens com base no filtro selecionado
        self.ax.clear()
        contagens_completo = [self.contagens.get(sector, 0) for sector in self.setores]
        self.ax.bar(self.setores, contagens_completo, color='green')
        self.ax.set_title(f"Contagem de BO's por setor ({self.ano_atual})")
        self.ax.set_xlabel("Setor")
        self.ax.set_ylabel("Contagem")

        formatter = FuncFormatter(lambda x, _: int(x))
        self.ax.yaxis.set_major_formatter(formatter)

        plt.xticks(rotation=45)
        self.canvas.draw()

    def obter_contagens_por_setor_por_ano(self, ano):
        conn = create_connection()
        cursor = conn.cursor()

        if ano == "Todos":
            cursor.execute(f"SELECT COALESCE(setor_responsavel, 'Não especificado'), COUNT(*) FROM bo_records WHERE modulo like ? GROUP BY setor_responsavel", (self.ultimo_modulo,))
        else:
            query = ("SELECT COALESCE(setor_responsavel, 'Não especificado'), COUNT(*) FROM bo_records WHERE YEAR(emissao_totvs) = ? AND modulo like ? GROUP BY setor_responsavel")
            cursor.execute(query, (ano, self.ultimo_modulo))
        contagens = {row[0]: row[1] for row in cursor.fetchall()}
        cursor.close()
        conn.close()

        return contagens   

class Relatorios(TreeViewHelper):
    def __init__(self, parent, caller_id=None):
        self.parent = parent
        self.root = tk.Toplevel(parent)
        self.root.title("Gerar Relatório")
        self.root.iconbitmap(resource_path('SAREM.ico'))
        self.root.geometry("800x500")
        self.root.state('zoomed')

        self.root.transient(parent)
        center_window_child(self.root, parent)

        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(1, weight=1)
        
        self.ultimo_modulo = caller_id
        print(self.identificar_chamador())

        self.setores = self.setor()
        self.motivos = self.motivo()
        self.clientes = self.cliente()

        frame_relat = tb.Labelframe(self.root, text="Critérios de Busca", padding=10)
        frame_relat.grid(row=0, column=0, sticky="n", padx=10, pady=10)

        frame_relat.grid_rowconfigure(0, weight=1)
        frame_relat.grid_columnconfigure(0, weight=1)

        tree_frame = tb.Frame(self.root)
        tree_frame.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)

        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # Lista de BOs Treeview
        self.tree = tb.Treeview(
            tree_frame,
            columns=("BO", "OP", "loja", "tipo_ocorrencia", "setor_responsavel", "motivo"),
            show="headings"
        )
        self.tree.heading("BO", text="BO")
        self.tree.heading("OP", text="OP")
        self.tree.heading("loja", text="Cliente")
        self.tree.heading("tipo_ocorrencia", text="Tipo de Ocorrência")
        self.tree.heading("setor_responsavel", text="Setor Responsável")
        self.tree.heading("motivo", text="Motivo")
        self.tree.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

        # Barra de rolagem vertical
        v_scrollbar = tb.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=v_scrollbar.set)

        self.tree.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        v_scrollbar.grid(row=0, column=1, sticky="ns")

        self.meses = ["Todos"] + [f"{i:02d}" for i in range(1, 13)]

        # Campos de Filtros

        # Clientes
        tk.Label(frame_relat, text="Cliente:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.clientes_combobox = tb.Combobox(frame_relat, values=self.clientes, state="readonly")
        self.clientes_combobox.grid(row=0, column=1, padx=5, pady=5, sticky="nsew")
        self.clientes_combobox.set("Todos")

        # Setores
        tk.Label(frame_relat, text="Setor:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        self.setores_combobox = tb.Combobox(frame_relat, values=self.setores, state="readonly")
        self.setores_combobox.grid(row=1, column=1, padx=5, pady=5)
        self.setores_combobox.set("Todos")

        # Motivos
        tk.Label(frame_relat, text="Motivo:").grid(row=2, column=0, sticky=tk.W, padx=5, pady=5)
        self.motivos_combobox = tb.Combobox(frame_relat, values=self.motivos, state="readonly")
        self.motivos_combobox.grid(row=2, column=1, padx=5, pady=5)
        self.motivos_combobox.set("Todos")

        # Meses
        tk.Label(frame_relat, text="Mês:").grid(row=3, column=0, sticky=tk.W, padx=5, pady=5)
        self.meses_combobox = tb.Combobox(frame_relat, values=self.meses, state="readonly")
        self.meses_combobox.grid(row=3, column=1, padx=5, pady=5)
        self.meses_combobox.set("Todos")

        # Botão Gerar Relatório
        tb.Button(frame_relat, text="Gerar Relatório", command=self.gerar_relatorios).grid(row=4, column=0, columnspan=2, pady=10)

        # Botões de Exportação
        tb.Button(frame_relat, text="Imprimir", command=self.imprimir_relatorio).grid(row=5, column=0, columnspan=2, pady=10)
        tb.Button(frame_relat, text="Salvar PDF", command=self.exportar_para_pdf).grid(row=6, column=0, columnspan=2, pady=10)
        tb.Button(frame_relat, text="Salvar Excel", command=self.exportar_para_excel).grid(row=7, column=0, columnspan=2, pady=10)

        
    def gerar_relatorios(self):
        conn = create_connection()
        cursor = conn.cursor()
        
        cliente = self.clientes_combobox.get()
        setor = self.setores_combobox.get()
        motivo = self.motivos_combobox.get()
        mes = self.meses_combobox.get()
        
        try:
            # Limpa os dados existentes na Treeview
            for item in self.tree.get_children():
                self.tree.delete(item)

            # Base da consulta SQL
            query = "SELECT * FROM bo_records WHERE modulo LIKE ? AND D_E_L_E_T_ <> '*' AND status NOT LIKE 'Embarcado'"
            params = [self.ultimo_modulo]

            # Adiciona condições dinamicamente com base nos filtros
            if cliente != "Todos":
                query += " AND loja = ?"
                params.append(cliente)
            if setor != "Todos":
                query += " AND setor_responsavel = ?"
                params.append(setor)
            if motivo != "Todos":
                query += " AND motivo = ?"
                params.append(motivo)
            if mes != "Todos":
                query += " AND MONTH(emissao_totvs) = ?"
                params.append(int(mes))

            # Executa a consulta
            cursor.execute(query, params)
            rows = cursor.fetchall()

            # Insere os resultados na Treeview
            for row in rows:
                bo = str(row[1]).strip() if row[1] is not None else ""
                op = str(row[2]).strip() if row[2] is not None else ""
                cliente = str(row[3]).strip() if row[3] is not None else ""
                tipo_ocorrencia = str(row[5]).strip() if row[5] is not None else ""
                setor_responsavel = str(row[8]).strip() if row[8] is not None else ""
                motivo = str(row[6]).strip() if row[6] is not None else ""

                self.tree.insert(
                    "",
                    "end",
                    values=(bo, op, cliente, tipo_ocorrencia, setor_responsavel, motivo)
                )
                
        except pyodbc.Error as e:
            messagebox.showerror("Erro", f"Erro ao carregar BOs: {e}")
        finally:
            if conn:
                cursor.close()
                conn.close()
        self.ajustar_colunas()

    def exportar_para_pdf(self, file_path=None):
        if file_path is None:
            # Abre uma janela para salvar o arquivo
            file_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
            )
            if not file_path:
                return

        # Configurações do documento
        doc = SimpleDocTemplate(
            file_path,
            pagesize=landscape(A4),
            leftMargin=1*cm,
            rightMargin=1*cm,
            topMargin=1*cm,
            bottomMargin=1*cm
        )
        elements = []

        # Estilo dos parágrafos
        styleSheet = getSampleStyleSheet()
        style = styleSheet['Normal']
        style.alignment = TA_LEFT
        style.fontSize = 10
        style.leading = 12

        # Obter dados da tabela
        headers = [self.tree.heading(col, "text") for col in self.tree["columns"]]
        data = [headers]
        
        for item in self.tree.get_children():
            values = list(self.tree.item(item, "values"))
            # Converter valores para parágrafos com quebra de linha
            formatted_values = []
            for value in values:
                # Limitar linha a 30 caracteres e quebrar texto
                wrapped_text = '\n'.join(wrap(str(value), 30))
                formatted_values.append(Paragraph(wrapped_text, style))
            data.append(formatted_values)

        # Criar tabela com larguras automáticas
        col_widths = [None] * len(headers)  # Larguras automáticas
        t = Table(data, colWidths=col_widths, repeatRows=1)

        # Estilo da tabela
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ])
        t.setStyle(style)

        elements.append(t)
        doc.build(elements)
        messagebox.showinfo("Sucesso", f"Relatório exportado para {file_path} com sucesso!")

    def exportar_para_excel(self, file_path=None):
        if file_path is None:
            # Abre uma janela para salvar o arquivo
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if not file_path:
                return
        
        try:
            # Extrai dados da Treeview
            headers = [self.tree.heading(col, "text") for col in self.tree["columns"]]
            data = []
            
            for item in self.tree.get_children():
                values = list(self.tree.item(item, "values"))
                formatted_values = []
                for value in values:
                    # Quebra texto em linhas de até 30 caracteres
                    wrapped_text = '\n'.join(wrap(str(value), 30))
                    formatted_values.append(wrapped_text)
                data.append(formatted_values)

            # Cria DataFrame e exporta para Excel
            df = pd.DataFrame(data, columns=headers)
            df.to_excel(file_path, index=False)

            # Carrega workbook para formatação avançada
            wb = load_workbook(file_path)
            ws = wb.active

            # Formatação de células
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
            row_fills = [PatternFill(start_color="FFFFFF", fill_type="solid"),
                        PatternFill(start_color="FFFFE0", fill_type="solid")]
            border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

            # Aplica formatação
            for row_idx, row in enumerate(ws.iter_rows(), start=1):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(wrap_text=True, vertical='center')
                    
                    if row_idx == 1:  # Cabeçalho
                        cell.font = header_font
                        cell.fill = header_fill
                    else:  # Linhas alternadas
                        cell.fill = row_fills[(row_idx - 2) % 2]

            # Ajusta largura das colunas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        lines = str(cell.value).split('\n')
                        current_length = max(len(line) for line in lines)
                        if current_length > max_length:
                            max_length = current_length
                ws.column_dimensions[column].width = max_length + 2

            # Salva as alterações
            wb.save(file_path)
            messagebox.showinfo("Sucesso", f"Arquivo exportado para:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao exportar:\n{str(e)}")

    def imprimir_relatorio(self):
        try:
            # Gera o PDF temporariamente
            temp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
            temp_pdf.close()
            
            # Exporta para o PDF temporário usando a função existente
            self.exportar_para_pdf(temp_pdf.name)
            
            # Abre a janela de impressão
            if sys.platform == "win32":
                os.startfile(temp_pdf.name, "print")
            else:
                # Para macOS e Linux
                subprocess.run(["lp", temp_pdf.name])
            
            # Aguarda um breve período para garantir que a impressão tenha sido processada
            time.sleep(5)
            
            # Remove o arquivo temporário após a impressão
            if os.path.exists(temp_pdf.name):
                os.unlink(temp_pdf.name)
            messagebox.showinfo("Sucesso", "Relatório enviado para impressão!")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao imprimir: {str(e)}")

    def cliente(self):
        conn = create_connection()
        cursor = conn.cursor()

        query = ("SELECT COALESCE(loja, 'Não especificado') AS Setor FROM bo_records WHERE loja IS NOT NULL AND loja NOT LIKE '' AND modulo LIKE ? AND D_E_L_E_T_ <> '*' GROUP BY loja")
        cursor.execute(query, (self.ultimo_modulo))
        cliente = ["Todos"] + [row[0] for row in cursor.fetchall()]

        cursor.close()
        conn.close()

        return cliente
    
    def setor(self):
        conn = create_connection()
        cursor = conn.cursor()

        query = ("SELECT COALESCE(setor_responsavel, 'Não especificado') AS Setor FROM bo_records WHERE setor_responsavel IS NOT NULL AND setor_responsavel NOT LIKE '' AND modulo LIKE ? AND D_E_L_E_T_ <> '*' GROUP BY setor_responsavel")
        cursor.execute(query, (self.ultimo_modulo))
        setores = ["Todos"] + [row[0] for row in cursor.fetchall()]

        cursor.close()
        conn.close()

        return setores

    def motivo(self):
        conn = create_connection()
        cursor = conn.cursor()

        query = ("SELECT COALESCE(motivo, 'Não especificado') AS Setor FROM bo_records WHERE motivo IS NOT NULL AND motivo NOT LIKE '' AND modulo LIKE ? AND D_E_L_E_T_ <> '*' GROUP BY motivo")
        cursor.execute(query, (self.ultimo_modulo))
        motivos = ["Todos"] + [row[0] for row in cursor.fetchall()]

        cursor.close()
        conn.close()

        return motivos

    def identificar_chamador(self):
        if self.ultimo_modulo is None:
            raise ValueError(
                "É necessário informar o identificador do módulo que chamou a função")
        return f"Tela de Relatórios chamada pelo módulo: {self.ultimo_modulo}"
    
class gerarBO():
    def __init__(self, parent, caller_id=None):
        self.parent = parent
        self.root = tk.Toplevel(parent)
        self.root.title("Solicitação de nova BO")
        self.root.iconbitmap(resource_path('SAREM.ico'))
        self.root.minsize(400, 300)
        self.root.resizable(False, False)
        self.root.focus_set()

        self.root.transient(parent)

        self.ultimo_modulo = caller_id
        print(self.identificar_chamador())

        campos = [
            ("cliente", "Cliente"),
            ("loja", "Loja"),
            ("n_de_bo", "Número de BO"),
            ("n_de_pedido", "Número de Pedido"),
            ("endereco", "Endereço"),
            ("invoice_inicial", "Invoice Inicial"),
            ("nf_de_fabrica", "NF de Fábrica"),
            ("data_emissao", "Data de Emissão"),
            ("nf_envio", "NF de Envio"),
        ]
        self.entries = {}

        frame = tb.Frame(self.root, padding=20)
        frame.pack(expand=True, fill=tk.BOTH)

        tb.Label(frame, text="Boletim de Ocorrência", font=("Arial", 10, "bold")).grid(row=0, column=0, columnspan=2, pady=(0, 10))
        for i, (campo, label) in enumerate(campos):
            tb.Label(frame, text=label + ":").grid(row=i+1, column=0, sticky=tk.W, pady=5)
            entry = tb.Entry(frame, width=30, show="*" if campo == "password" else "")
            entry.grid(row=i+1, column=1, pady=5)
            self.entries[campo] = entry

        # row_btn = len(campos) + 2
        # tb.Button(frame, text="Salvar", command=self.salvar).grid(row=row_btn, column=0, pady=15, sticky="ew")
        # tb.Button(frame, text="Cancelar", command=self.cancelar).grid(row=row_btn, column=1, pady=15, sticky="ew")

        self.root.update_idletasks()
        largura = self.root.winfo_reqwidth()
        altura = self.root.winfo_reqheight()
        self.root.geometry(f"{largura}x{altura}")
        center_window_child(self.root, parent)

    def identificar_chamador(self):
        if self.ultimo_modulo is None:
            raise ValueError(
                "É necessário informar o identificador do módulo que chamou a função")
        return f"Tela de Gerar BO chamada pelo módulo: {self.ultimo_modulo}"

class buscarBo(TreeViewHelper):
    def __init__(self, parent, caller_id=None, user=None):
        self.user = user
        self.parent = parent
        self.root = tk.Toplevel(parent)
        self.root.title("Acompanhar uma BO")
        self.root.iconbitmap(resource_path('SAREM.ico'))
        self.root.geometry("1000x600")
        self.root.state('zoomed')

        self.root.transient(parent)

        self.ultimo_modulo = caller_id

        from theme import ajustar_largura_colunas

        # Cabeçalho
        header = tb.Frame(self.root)
        header.pack(fill=tk.X, padx=10, pady=10)

        tb.Button(header, text="Fechar",
                   command=lambda: self.root.destroy()).pack(side=tk.RIGHT)

        # Barra de pesquisa
        self.search_bar = SearchBar(
            self.root, self.pesquisar_bo, self.clear_search)

        # Frame para Treeview e barra de rolagem
        tree_frame = tb.Frame(self.root)
        tree_frame.pack(expand=True, fill=tk.BOTH, padx=10, pady=(0, 10))

        # lista de BOs Treeview
        self.tree = tb.Treeview(
            tree_frame,
            columns=("C5_PEDREPR", "C5_NUM", "C5_NOME",
                     "C5_FILIAL", "C5_EMISSAO", "C5_ENTREGA"),
            show="headings"
        )
        self.tree.heading("C5_PEDREPR", text="BO")
        self.tree.heading("C5_NUM", text="OP")
        self.tree.heading("C5_NOME", text="CLIENTE")
        self.tree.heading("C5_FILIAL", text="EMPRESA")
        self.tree.heading("C5_EMISSAO", text="EMISSÃO")
        self.tree.heading("C5_ENTREGA", text="PREVISÃO DE ENTREGA")

        # Barra de rolagem vertical
        v_scrollbar = tb.Scrollbar(
            tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.tree.configure(yscrollcommand=v_scrollbar.set)
        self.tree.pack(expand=True, fill=tk.BOTH, side=tk.LEFT)

        self.tree.bind("<Double-1>", lambda event: exibir_detalhes(self.root,
                       self.tree, caller_id=self.ultimo_modulo, user=self.user))

        center_window(self)
        self.carregar_bos()
        self.search_bar.search_entry.bind('<Return>', self.pesquisar_bo)


    def identificar_chamador(self):
        if self.ultimo_modulo is None:
            raise ValueError(
                "É necessário informar o identificador do módulo que chamou a função")
        return f"Tela de Buscar BO chamada pelo módulo: {self.ultimo_modulo}"

    def carregar_bos(self):
        conn = create_connection_Protheus()
        if conn is None:
            return

        try:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT C5_PEDREPR, C5_NUM, C5_NOME, IIF(C5_FILIAL='0101','TIDELLI','NAUTICA') AS FILIAL, CONVERT(VARCHAR,CONVERT(DATETIME,C5_EMISSAO),103) as EMISSAO, CONVERT(VARCHAR,CONVERT(DATETIME,C5_ENTREGA),103) as PREVISAO_ENTREGA
                FROM SC5010 SC5
                WHERE SC5.D_E_L_E_T_ <> '*'
                        AND SC5.C5_NOTA = ''
                        AND PATINDEX('BO[0-9]%', SC5.C5_PEDREPR) > 0
                        AND C5_FILIAL IN ('0101','0201')
                ORDER BY C5_EMISSAO DESC
            """)
            rows = cursor.fetchall()

            # Remove todos os itens atuais da Treeview
            for item in self.tree.get_children():
                self.tree.delete(item)

            # Insere os dados na Treeview
            for row in rows:
                bo = str(row[0]).strip() if row[0] is not None else ""
                op = str(row[1]).strip() if row[1] is not None else ""
                cliente = str(row[2]).strip() if row[2] is not None else ""
                filial = str(row[3]).strip() if row[3] is not None else ""
                emissao = str(row[4]).strip() if row[4] is not None else ""
                previsao_entrega = str(
                    row[5]).strip() if row[5] is not None else ""

                self.tree.insert("", tk.END, values=(
                    bo, op, cliente, filial, emissao, previsao_entrega))
                
        except pyodbc.Error as e:
            messagebox.showerror("Erro", f"Erro ao carregar BOs: {e}")
        finally:
            if conn:
                cursor.close()
                conn.close()
        ajustar_largura_colunas(self.tree)

    def pesquisar_bo(self, event=None):
        termo = self.search_bar.search_entry.get()
        if not termo:
            return

        conn = create_connection_Protheus()
        if conn is None:
            messagebox.showerror("Erro", "Falha ao conectar ao banco de dados")
            return

        try:
            cursor = conn.cursor()
            query = """
                SELECT C5_PEDREPR, C5_NUM, C5_NOME, IIF(C5_FILIAL='0101','TIDELLI','NAUTICA') AS FILIAL, CONVERT(VARCHAR,CONVERT(DATETIME,C5_EMISSAO),103) AS EMISSAO, CONVERT(VARCHAR,CONVERT(DATETIME,C5_ENTREGA),103) AS PREVISAO_ENTREGA
                FROM SC5010 SC5
                WHERE SC5.D_E_L_E_T_ <> '*'
                    AND SC5.C5_NOTA = ''
                    AND C5_PEDREPR LIKE 'BO%'  -- Garante que o campo começa com "BO"
                    AND (C5_PEDREPR LIKE 'BO%' + ?  -- Permite que o termo apareça após o "BO"
                        OR UPPER(C5_NUM) LIKE UPPER(?)
                        OR UPPER(C5_NOME) LIKE UPPER(?))
                    AND C5_FILIAL IN ('0101','0201')
                ORDER BY C5_EMISSAO DESC
                """

            cursor.execute(
                query, (f"%{termo}%", f"%{termo}%", f"%{termo}%"))
            rows = cursor.fetchall()

            self.tree.delete(*self.tree.get_children())
            for row in rows:
                bo = str(row[0]).strip() if row[0] is not None else ""
                op = str(row[1]).strip() if row[1] is not None else ""
                cliente = str(row[2]).strip() if row[2] is not None else ""
                filial = str(row[3]).strip() if row[3] is not None else ""
                emissao = str(row[4]).strip() if row[4] is not None else ""
                previsao_entrega = str(
                    row[5]).strip() if row[5] is not None else ""

                self.tree.insert("", tk.END, values=(
                    bo, op, cliente, filial, emissao, previsao_entrega))
                
            self.ajustar_colunas()
        except pyodbc.Error as e:
            messagebox.showerror("Erro", f"Erro ao pesquisar BOs: {e}")
        finally:
            if conn:
                cursor.close()
                conn.close()

    def clear_search(self):
        self.search_bar.search_entry.delete(0, tk.END)
        self.search_bar.update_buttons()  # Atualiza os botões após limpar
        self.carregar_bos()  # Recarrega os BOs

    def update_buttons(self, event=None):
        termo = self.search_entry.get()
        if termo:
            self.search_button.pack_forget()  # Oculta o botão "Pesquisar"
            self.clear_button.pack(side=tk.LEFT)  # Exibe o botão "Limpar"
        else:
            self.clear_button.pack_forget()  # Oculta o botão "Limpar"
            self.search_button.pack(side=tk.LEFT)


class exibir_detalhes():
    instance = None

    def __init__(self, parent, tree, caller_id=None, user=None):
        self.user = user
        self.root = tk.Toplevel(parent)
        self.root.title("Detalhes BO")
        self.root.iconbitmap(resource_path('SAREM.ico'))
        self.root.minsize(400, 300)
        self.root.resizable(False, False)
        self.root.focus_set()

        self.root.transient(parent)
        self.root.grab_set()

        self.tree = tree
        self.ultimo_modulo = caller_id

        exibir_detalhes.instance = self

        # Verifica se a BO já existe no banco de dados
        self.bo_ja_acompanhada = self.verificar_bo_acompanhada()

        # Se já está sendo acompanhada, exibe aviso no topo
        if self.bo_ja_acompanhada:
            aviso_frame = tb.Frame(self.root)
            aviso_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 0))
            aviso_label = tb.Label(
                aviso_frame,
                text="⚠️ Esta BO já está sendo acompanhada!",
                foreground="red",
                font=("Arial", 10, "bold")
            )
            aviso_label.pack(fill="x")

            detalhe_row = 1
        else:
            detalhe_row = 0

        self.sc5_detalhes(row=detalhe_row)
        self.itens_bo(row=detalhe_row + 1)
        self.opcoes_bo(row=detalhe_row + 2)

        self.root.update_idletasks()
        center_window_child(self.root, parent)

    def verificar_bo_acompanhada(self):
        item_selecionado = self.tree.selection()
        if not item_selecionado:
            return False
        valores = self.tree.item(item_selecionado, "values")
        if not valores or not valores[0]:
            return False
        bo_number = valores[0]
        conn = create_connection()
        if conn is None:
            return False
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT 1 FROM bo_records WHERE bo_number = ? AND D_E_L_E_T_ <> '*'", (bo_number,))
            exists = cursor.fetchone()
            return exists is not None
        except Exception:
            return False
        finally:
            if conn:
                cursor.close()
                conn.close()

    def sc5_detalhes(self, row=0):
        self.root.grid_rowconfigure(row, weight=3)
        self.root.grid_columnconfigure(0, weight=1)

        frame_sc5Detalhes = tb.LabelFrame(
            self.root, text="Detalhes da Ocorrência", padding=10)
        frame_sc5Detalhes.grid(
            row=row, column=0, sticky="nsew", padx=10, pady=10)

        frame_sc5Detalhes.grid_rowconfigure(0, weight=1)
        frame_sc5Detalhes.grid_columnconfigure(0, weight=1)
        frame_sc5Detalhes.grid_columnconfigure(1, weight=2)

        # Obtém o item selecionado no Treeview
        item_selecionado = self.tree.selection()
        if not item_selecionado:
            return

        # Obtém os valores do item selecionado
        valores = self.tree.item(item_selecionado, "values")

        # Criando as labels
        labels = ["BO:", "OP:", "CLIENTE:", "FILIAL:",
                  "EMISSÃO:", "PREVISÃO DE ENTREGA:"]
        for i, label_text in enumerate(labels):
            tb.Label(frame_sc5Detalhes, text=label_text).grid(
                row=i, column=0, sticky="w", padx=5, pady=2)
            tb.Label(frame_sc5Detalhes, text=valores[i]).grid(
                row=i, column=1, sticky="ew", padx=5, pady=2)

        self.root.update_idletasks()
        self.root.minsize(400, self.root.winfo_height())

    def itens_bo(self, row=1):
        self.root.grid_rowconfigure(row, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

        frame_itensBo = tb.LabelFrame(
            self.root, text="Itens da BO", padding=10)
        frame_itensBo.grid(row=row, column=0, sticky="nsew", padx=10, pady=10)

        frame_itensBo.grid_rowconfigure(0, weight=1)
        for i in range(3):
            frame_itensBo.grid_columnconfigure(i, weight=1)

        labels = ["CÓDIGO", "DESCRIÇÃO", "LINHA"]

        item_selecionado2 = self.tree.selection()
        if not item_selecionado2:
            return

        valores2 = self.tree.item(item_selecionado2, "values")

        conn = create_connection_Protheus()
        if conn is None:
            messagebox.showerror("Erro", "Falha ao conectar ao banco de dados")
            return

        cursor = conn.cursor()

        query = """SELECT SC6.C6_CODTIDI, SC6.C6_DESCRI, SC6.C6_LINHA
        FROM SC6010 SC6
        INNER JOIN SC5010 SC5 ON (SC5.C5_NUM = SC6.C6_NUM AND SC5.C5_FILIAL = SC6.C6_FILIAL AND SC5.D_E_L_E_T_ <> '*')
        WHERE SC6.C6_NUM LIKE ? AND SC5.C5_PEDREPR LIKE ? AND SC5.C5_PEDREPR LIKE ?"""

        try:
            cursor.execute(
                query, (f"%{valores2[1]}%", "%BO%", f"%{valores2[0]}%"))
            itens_bo = cursor.fetchall()
        except pyodbc.Error as e:
            messagebox.showerror("Erro", f"Erro ao executar a consulta: {e}")
            return
        finally:
            cursor.close()
            conn.close()

        for i, label_txt in enumerate(labels):
            label = tb.Label(frame_itensBo, text=label_txt)
            label.grid(row=0, column=i, sticky="w", padx=5, pady=2)
            frame_itensBo.grid_columnconfigure(i, weight=1)

        for idx, item in enumerate(itens_bo):
            for col, value in enumerate(item):
                label = tb.Label(frame_itensBo, text=value)
                label.grid(row=idx + 1, column=col, sticky="w", padx=5, pady=2)
                frame_itensBo.grid_columnconfigure(col, weight=1)

        self.root.update_idletasks()
        self.root.geometry("")

    def opcoes_bo(self, row=2):
        self.root.grid_rowconfigure(row, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

        frame_opcoesBo = tb.LabelFrame(
            self.root, text="Opções", padding=10)
        frame_opcoesBo.grid(row=row, column=0, sticky="nsew", padx=10, pady=10)

        frame_opcoesBo.grid_rowconfigure(0, weight=1)
        frame_opcoesBo.grid_columnconfigure(2, weight=1)

        tb.Button(frame_opcoesBo, text="Acompanhar BO", command=self.acompanhar_bo).grid(
            row=0, column=0, sticky="e", padx=5, pady=2)

    def acompanhar_bo(self):
        obj = acompanhar_Bo(self.root, self.tree,
                            caller_id=self.ultimo_modulo, user=self.user)
        resultado = obj.identificar_chamador()
        print(resultado)

class exibir_detalhes_acompanhando():
    instance = None

    def __init__(self, parent, tree, caller_id=None, user=None):
        self.parent = parent
        self.user = user
        self.root = tk.Toplevel(parent)
        self.root.title("Detalhes BO")
        self.root.iconbitmap(resource_path('SAREM.ico'))
        self.root.minsize(400, 300)
        self.root.resizable(False, False)
        self.root.focus_set()

        self.root.transient(parent)
        self.root.grab_set()

        self.tree = tree
        self.ultimo_modulo = caller_id

        exibir_detalhes_acompanhando.instance = self

        self.sc5_detalhes()
        self.itens_bo()
        self.opcoes_bo()

        self.root.update_idletasks()
        center_window_child(self.root, parent)

    def sc5_detalhes(self):
        self.root.grid_rowconfigure(0, weight=3)
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

        frame_sc5Detalhes = tb.LabelFrame(
            self.root, text="Detalhes da Ocorrência", padding=10)
        frame_sc5Detalhes.grid(
            row=0, column=0, sticky="nsew", padx=10, pady=10)

        frame_sc5Detalhes.grid_rowconfigure(0, weight=1)
        frame_sc5Detalhes.grid_columnconfigure(0, weight=1)
        frame_sc5Detalhes.grid_columnconfigure(1, weight=2)

        # Obtém o item selecionado no Treeview
        item_selecionado = self.tree.selection()
        if not item_selecionado:
            return

        # Obtém os valores do item selecionado
        valores = self.tree.item(item_selecionado, "values")

        conn = create_connection()
        if conn is None:
            messagebox.showerror("Erro", "Falha ao conectar ao banco de dados")
            return

        cursor = conn.cursor()

        query = """SELECT bo_number, op, loja, tipo_ocorrencia, CONVERT(VARCHAR,CONVERT(DATETIME,previsao_embarque),103), descricao, frete, [status]
        FROM bo_records
        WHERE bo_number LIKE ?"""

        try:
            cursor.execute(
                query, (f"%{valores[0]}%"))
            itens_bo = cursor.fetchall()
        except pyodbc.Error as e:
            messagebox.showerror("Erro", f"Erro ao executar a consulta: {e}")
            return
        finally:
            cursor.close()
            conn.close()

        # Criando as labels
        labels = ["BO:", "OP:", "CLIENTE:", "TIPO DE OCORRÊNCIA:", "PREVISÃO DE EMBARQUE:", "DESCRIÇÃO:", "FRETE:", "STATUS:"]
        if itens_bo:
            for i, label_text in enumerate(labels):
                tb.Label(frame_sc5Detalhes, text=label_text).grid(
                    row=i, column=0, sticky="w", padx=5, pady=2)
                tb.Label(frame_sc5Detalhes, text=itens_bo[0][i]).grid(
                    row=i, column=1, sticky="ew", padx=5, pady=2)
        else:
            messagebox.showinfo("Aviso", "Nenhum detalhe encontrado para este BO.")

        self.root.update_idletasks()
        self.root.minsize(400, self.root.winfo_height())

    def itens_bo(self):
        self.root.grid_rowconfigure(0, weight=3)
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

        frame_itensBo = tb.LabelFrame(
            self.root, text="Itens da BO", padding=10)
        frame_itensBo.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)

        frame_itensBo.grid_rowconfigure(0, weight=1)
        for i in range(3):
            frame_itensBo.grid_columnconfigure(i, weight=1)

        labels = ["CÓDIGO", "DESCRIÇÃO", "LINHA", "MOTIVO"]

        item_selecionado2 = self.tree.selection()
        if not item_selecionado2:
            return

        valores2 = self.tree.item(item_selecionado2, "values")

        conn = create_connection()
        if conn is None:
            messagebox.showerror("Erro", "Falha ao conectar ao banco de dados")
            return

        cursor = conn.cursor()

        query = """SELECT COD, [DESC], LINHA, BI.MOTIVO
        FROM BO_ITENS as BI
        INNER JOIN bo_records AS BR ON BR.bo_number = BI.BO_REF
        WHERE BR.bo_number LIKE ? AND BR.op LIKE ? AND BR.D_E_L_E_T_ <> '*'"""

        try:
            cursor.execute(
                query, (f"%{valores2[0]}%", f"%{valores2[1]}%"))
            itens_bo = cursor.fetchall()
        except pyodbc.Error as e:
            messagebox.showerror("Erro", f"Erro ao executar a consulta: {e}")
            return
        finally:
            cursor.close()
            conn.close()

        for i, label_txt in enumerate(labels):
            label = tb.Label(frame_itensBo, text=label_txt)
            label.grid(row=0, column=i, sticky="w", padx=5, pady=2)
            frame_itensBo.grid_columnconfigure(i, weight=1)

        for idx, item in enumerate(itens_bo):
            for col, value in enumerate(item):
                label = tb.Label(frame_itensBo, text=value)
                label.grid(row=idx + 1, column=col, sticky="w", padx=5, pady=2)
                frame_itensBo.grid_columnconfigure(col, weight=1)

        self.root.update_idletasks()
        self.root.geometry("")

    def opcoes_bo(self, row=2):
        self.root.grid_rowconfigure(row, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

        frame_opcoesBo = tb.LabelFrame(
            self.root, text="Opções", padding=10)
        frame_opcoesBo.grid(row=row, column=0, sticky="nsew", padx=10, pady=10)

        frame_opcoesBo.grid_rowconfigure(0, weight=1)
        frame_opcoesBo.grid_columnconfigure(2, weight=1)

        tb.Button(frame_opcoesBo, text="Editar BO", command=self.editar_bo).grid(
            row=0, column=0, sticky="e", padx=5, pady=2)
        tb.Button(frame_opcoesBo, text="Excluir BO", command=self.excluir_bo).grid(
            row=0, column=1, sticky="e", padx=5, pady=2)

    def editar_bo(self):
        self.root.withdraw()
        editor = editar_Bo(self.parent, self.tree, caller_id=self.ultimo_modulo, user=self.user)
        self.root.wait_window(editor.root)
        try:
            self.root.deiconify()
        except tk.TclError:
            pass
        

    def excluir_bo(self):
        print(f"Usuário {self.user[1]} solicitou a exclusão da BO.")

        resposta = messagebox.askyesno("Confirmação",
            "Você tem certeza que deseja excluir esta BO? Esta ação não pode ser desfeita.",
            icon=messagebox.WARNING
        )

        item_selecionado = self.tree.selection()
        if not item_selecionado:
            messagebox.showwarning("Aviso", "Nenhum item selecionado.")
            return
        
        valores = self.tree.item(item_selecionado, "values")
        bo_number = valores[0]

        
        if resposta is True:
            try:
                conn = create_connection()
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE bo_records
                    SET D_E_L_E_T_ = '*', user_delet = ?, deleted_at = GETDATE()
                    WHERE bo_number = ?
                """, (self.user[1]), str(bo_number,))
                conn.commit()
                messagebox.showinfo("Sucesso", f"{bo_number} excluída com sucesso!")
                self.root.destroy()
            except pyodbc.Error as e:
                messagebox.showerror("Erro", f"Erro ao excluir BO: {e}")
                self.root.destroy()
            finally:
                if conn:
                    cursor.close()
                    conn.close()

            if self.ultimo_modulo == "Corporativo":
                from modulos.corporativo import CorporativoModule
                if CorporativoModule.instance is not None:
                    if CorporativoModule.instance is not None:
                        CorporativoModule.instance.atualizar_bos()
                    else:
                        print("Instância do módulo corporativo não encontrada.")
                else:
                    pass
            elif self.ultimo_modulo == "Varejo":
                from modulos.varejo import VarejoModule
                if VarejoModule.instance is not None:
                    if VarejoModule.instance is not None:
                        VarejoModule.instance.atualizar_bos()
                    else:
                        print("Instância do módulo varejo não encontrada.")
                else:
                    pass
            elif self.ultimo_modulo == "Exportacao":
                from modulos.exportacao import ExportacaoModule
                if ExportacaoModule.instance is not None:
                    if ExportacaoModule.instance is not None:
                        ExportacaoModule.instance.atualizar_bos()
                    else:
                        print("Instância do módulo exportacao não encontrada.")
                else:
                    pass
        else:
            pass

class exibir_detalhes_embarcado():
    instance = None

    def __init__(self, parent, tree, caller_id=None, user=None):
        self.parent = parent
        self.user = user
        self.root = tk.Toplevel(parent)
        self.root.title("Detalhes BO")
        self.root.iconbitmap(resource_path('SAREM.ico'))
        self.root.minsize(400, 300)
        self.root.resizable(False, False)
        self.root.focus_set()

        self.root.transient(parent)
        self.root.grab_set()

        self.tree = tree
        self.ultimo_modulo = caller_id

        exibir_detalhes_acompanhando.instance = self

        self.sc5_detalhes()
        self.itens_bo()

        self.root.update_idletasks()
        center_window_child(self.root, parent)

    def sc5_detalhes(self):
        self.root.grid_rowconfigure(0, weight=3)
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

        frame_sc5Detalhes = tb.LabelFrame(
            self.root, text="Detalhes da Ocorrência", padding=10)
        frame_sc5Detalhes.grid(
            row=0, column=0, sticky="nsew", padx=10, pady=10)

        frame_sc5Detalhes.grid_rowconfigure(0, weight=1)
        frame_sc5Detalhes.grid_columnconfigure(0, weight=1)
        frame_sc5Detalhes.grid_columnconfigure(1, weight=2)

        # Obtém o item selecionado no Treeview
        item_selecionado = self.tree.selection()
        if not item_selecionado:
            return

        # Obtém os valores do item selecionado
        valores = self.tree.item(item_selecionado, "values")

        conn = create_connection()
        if conn is None:
            messagebox.showerror("Erro", "Falha ao conectar ao banco de dados")
            return

        cursor = conn.cursor()

        query = """SELECT bo_number, op, loja, tipo_ocorrencia, CONVERT(VARCHAR,CONVERT(DATETIME,previsao_embarque),103), descricao, frete, [status]
        FROM bo_records
        WHERE bo_number LIKE ?"""

        try:
            cursor.execute(
                query, (f"%{valores[0]}%"))
            itens_bo = cursor.fetchall()
        except pyodbc.Error as e:
            messagebox.showerror("Erro", f"Erro ao executar a consulta: {e}")
            return
        finally:
            cursor.close()
            conn.close()

        # Criando as labels
        labels = ["BO:", "OP:", "CLIENTE:", "TIPO DE OCORRÊNCIA:", "PREVISÃO DE EMBARQUE:", "DESCRIÇÃO:", "FRETE:", "STATUS:"]
        if itens_bo:
            for i, label_text in enumerate(labels):
                tb.Label(frame_sc5Detalhes, text=label_text).grid(
                    row=i, column=0, sticky="w", padx=5, pady=2)
                tb.Label(frame_sc5Detalhes, text=itens_bo[0][i]).grid(
                    row=i, column=1, sticky="ew", padx=5, pady=2)
        else:
            messagebox.showinfo("Aviso", "Nenhum detalhe encontrado para este BO.")

        self.root.update_idletasks()
        self.root.minsize(400, self.root.winfo_height())

    def itens_bo(self):
        self.root.grid_rowconfigure(0, weight=3)
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

        frame_itensBo = tb.LabelFrame(
            self.root, text="Itens da BO", padding=10)
        frame_itensBo.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)

        frame_itensBo.grid_rowconfigure(0, weight=1)
        for i in range(3):
            frame_itensBo.grid_columnconfigure(i, weight=1)

        labels = ["CÓDIGO", "DESCRIÇÃO", "LINHA", "MOTIVO"]

        item_selecionado2 = self.tree.selection()
        if not item_selecionado2:
            return

        valores2 = self.tree.item(item_selecionado2, "values")

        conn = create_connection()
        if conn is None:
            messagebox.showerror("Erro", "Falha ao conectar ao banco de dados")
            return

        cursor = conn.cursor()

        query = """SELECT COD, [DESC], LINHA, BI.MOTIVO
        FROM BO_ITENS as BI
        INNER JOIN bo_records AS BR ON BR.bo_number = BI.BO_REF
        WHERE BR.bo_number LIKE ? AND BR.op LIKE ? AND BR.D_E_L_E_T_ <> '*'"""

        try:
            cursor.execute(
                query, (f"%{valores2[0]}%", f"%{valores2[1]}%"))
            itens_bo = cursor.fetchall()
        except pyodbc.Error as e:
            messagebox.showerror("Erro", f"Erro ao executar a consulta: {e}")
            return
        finally:
            cursor.close()
            conn.close()

        for i, label_txt in enumerate(labels):
            label = tb.Label(frame_itensBo, text=label_txt)
            label.grid(row=0, column=i, sticky="w", padx=5, pady=2)
            frame_itensBo.grid_columnconfigure(i, weight=1)

        for idx, item in enumerate(itens_bo):
            for col, value in enumerate(item):
                label = tb.Label(frame_itensBo, text=value)
                label.grid(row=idx + 1, column=col, sticky="w", padx=5, pady=2)
                frame_itensBo.grid_columnconfigure(col, weight=1)

        self.root.update_idletasks()
        self.root.geometry("")


class editar_Bo:
    def __init__(self, parent, tree, caller_id=None, user=None):
        self.user = user
        self.tree = tree
        self.ultimo_modulo = caller_id

        # ---------- JANELA ----------
        self.root = tk.Toplevel(parent)
        self.root.title("Editar BO")
        self.root.iconbitmap(resource_path('SAREM.ico'))
        # self.root.minsize(420, 300)
        self.root.transient(parent)
        self.root.grab_set()
        self.root.focus_set()

        # ---------- SCROLL GLOBAL (parte SUPERIOR que expande) ----------
        canvas = tk.Canvas(self.root)
        v_scroll = tb.Scrollbar(self.root, orient="vertical", command=canvas.yview)
        v_scroll.pack(side="right", fill="y")

        # 1) pack canvas no topo (expande)
        canvas.pack(side="top", fill="both", expand=True)
        canvas.configure(yscrollcommand=v_scroll.set)

        scrollable_frame = tb.Frame(canvas)
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

        # ---------- CONTEÚDO ROLÁVEL ----------
        item_selecionado = tree.selection()
        if not item_selecionado:
            messagebox.showwarning("Aviso", "Nenhum item selecionado.")
            self.root.destroy()
            return

        valores = tree.item(item_selecionado, "values")
        self.bo_number = valores[0]

        self.sc5_detalhes(valores, scrollable_frame)
        self.editaveis(parent, scrollable_frame)
        self.itens_bo(scrollable_frame)

        # ---------- SCROLL COM A RODA DO MOUSE ----------
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind_all("<MouseWheel>", _on_mousewheel)   # Windows / macOS

        # ---------- FRAME DE OPÇÕES FIXO EM BAIXO ----------
        self.opcoes_editor(self.root)          # empacota no root

        # ---------- AJUSTE DE TAMANHO ----------
        self.root.update_idletasks()
        ideal_width = scrollable_frame.winfo_reqwidth() + v_scroll.winfo_width() + 40
        self.root.geometry(f"{ideal_width}x600")
        center_window_child(self.root, parent)

        # centraliza
        self.root.update_idletasks()
        center_window_child(self.root, parent)

    def sc5_detalhes(self, valores, parent_frame):
        parent_frame.grid_rowconfigure(0, weight=3)
        parent_frame.grid_rowconfigure(0, weight=1)
        parent_frame.grid_columnconfigure(0, weight=1)

        frame_sc5Detalhes = tb.LabelFrame(
            parent_frame, text="Detalhes da Ocorrência", padding=10)
        frame_sc5Detalhes.grid(
            row=0, column=0, sticky="nsew", padx=10, pady=10)

        frame_sc5Detalhes.grid_rowconfigure(0, weight=1)
        frame_sc5Detalhes.grid_columnconfigure(0, weight=1)
        frame_sc5Detalhes.grid_columnconfigure(1, weight=2)

        # Obtém o item selecionado no Treeview
        item_selecionado = self.tree.selection()
        if not item_selecionado:
            return

        # Obtém os valores do item selecionado
        valores = self.tree.item(item_selecionado, "values")

        conn = create_connection()
        if conn is None:
            messagebox.showerror("Erro", "Falha ao conectar ao banco de dados")
            return

        cursor = conn.cursor()

        query = """SELECT bo_number, op, loja, CONVERT(VARCHAR,CONVERT(DATETIME,previsao_embarque),103), [status]
        FROM bo_records
        WHERE bo_number LIKE ?"""

        try:
            cursor.execute(
                query, (f"%{valores[0]}%",))
            itens_bo = cursor.fetchall()
        except pyodbc.Error as e:
            messagebox.showerror("Erro", f"Erro ao executar a consulta: {e}")
            return
        finally:
            cursor.close()
            conn.close()

        # Criando as labels
        labels = ["BO:", "OP:", "CLIENTE:", "PREVISÃO DE EMBARQUE:", "STATUS:"]
        if itens_bo:
            for i, label_text in enumerate(labels):
                tb.Label(frame_sc5Detalhes, text=label_text).grid(
                    row=i, column=0, sticky="w", padx=5, pady=2)
                tb.Label(frame_sc5Detalhes, text=itens_bo[0][i]).grid(
                    row=i, column=1, sticky="ew", padx=5, pady=2)
        else:
            messagebox.showinfo("Aviso", "Nenhum detalhe encontrado para este BO.")

        self.root.update_idletasks()
        self.root.minsize(400, self.root.winfo_height())

    def editaveis(self, parent, parent_frame):
        parent_frame.grid_rowconfigure(0, weight=1)

        # Frame para campos editáveis
        self.entries = {}
        campos = [
            ("tipo_ocorrencia", "Tipo de Ocorrência"),
            ("descricao", "Descrição"),
            ("setor_responsavel", "Setor Responsável"),
            ("frete", "Frete"),
        ]

        frame_editaveis = tb.LabelFrame(parent_frame, text="Campos Editáveis", padding=20)
        frame_editaveis.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)

        try:
            conn = create_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM bo_records WHERE bo_number = ?", (self.bo_number,))
            self.valores = cursor.fetchone()
            if self.valores is None:
                messagebox.showerror("Erro", f"BO {self.bo_number} não encontrada.")
                self.root.destroy()
                return
            cursor.close()
            conn.close()
        except pyodbc.Error as e:
            messagebox.showerror("Erro", f"Erro ao carregar BOs: {e}")
            self.root.destroy()
            return

        # Índices dos campos na tupla retornada pelo banco
        valor_idx = {
            "tipo_ocorrencia": 5,
            "motivo": 6,
            "descricao": 11,
            "setor_responsavel": 8,
            "frete": 7
        }

        for i, (campo, label_text) in enumerate(campos):
            tb.Label(frame_editaveis, text=label_text + ":").grid(
                row=i, column=0, sticky=tk.W, pady=5)
            if campo == "frete":
                entry = tb.Combobox(frame_editaveis, values=["CIF", "FOB"], state="readonly")
            elif campo == "setor_responsavel":
                entry = tb.Combobox(frame_editaveis, values=self.obter_setores(), state="readonly")
            elif campo == "descricao":
                entry = tk.Text(frame_editaveis, width=30, height=5, font=("Arial", 8))
            else:
                entry = tb.Entry(frame_editaveis, width=30)
            idx = valor_idx.get(campo)
            if idx is not None and self.valores and len(self.valores) > idx:
                valor = self.valores[idx]
                if campo == "frete":
                    entry.set(valor if valor else "")
                elif campo == "descricao":
                    entry.insert("1.0", valor if valor else "")
                else:
                    if isinstance(entry, tb.Combobox):
                        entry.set(valor if valor else "")
                    elif isinstance(entry, tb.Entry):
                        entry.insert(0, valor if valor else "")
            entry.grid(row=i, column=1, pady=5, sticky="ew")
            self.entries[campo] = entry

    def itens_bo(self, parent_frame):
        parent_frame.grid_rowconfigure(0, weight=3)
        parent_frame.grid_rowconfigure(0, weight=1)
        parent_frame.grid_columnconfigure(0, weight=1)

        frame_itensBo = tb.LabelFrame(
            parent_frame, text="Itens da BO", padding=10)
        frame_itensBo.grid(row=2, column=0, sticky="nsew", padx=10, pady=10)

        frame_itensBo.grid_rowconfigure(0, weight=1)
        for i in range(3):
            frame_itensBo.grid_columnconfigure(i, weight=1)

        labels = ["CÓDIGO", "DESCRIÇÃO", "LINHA", "MOTIVO"]

        item_selecionado2 = self.tree.selection()
        if not item_selecionado2:
            return

        valores2 = self.tree.item(item_selecionado2, "values")

        conn = create_connection()
        if conn is None:
            messagebox.showerror("Erro", "Falha ao conectar ao banco de dados")
            return

        cursor = conn.cursor()

        query = """SELECT COD, [DESC], LINHA, BI.MOTIVO
        FROM BO_ITENS as BI
        INNER JOIN bo_records AS BR ON BR.bo_number = BI.BO_REF
        WHERE BR.bo_number = ? AND BR.op = ? AND BR.D_E_L_E_T_ <> '*'"""

        try:
            cursor.execute(
                query, (f"{valores2[0]}", f"{valores2[1]}"))
            self.itens_bo = cursor.fetchall()
        except pyodbc.Error as e:
            messagebox.showerror("Erro", f"Erro ao executar a consulta: {e}")
            self.itens_bo = []
        finally:
            cursor.close()
            conn.close()

        # Cabeçalho
        for i, label_txt in enumerate(labels):
            label = tb.Label(frame_itensBo, text=label_txt)
            label.grid(row=0, column=i, sticky="w", padx=5, pady=2)
            frame_itensBo.grid_columnconfigure(i, weight=1)

        # Para guardar os widgets de motivo
        self.entries_motivo_itens = []

        # Linhas dos itens
        motivos_opcoes = self.obter_motivos()
        for idx, (cod, desc, linha, motivo) in enumerate(self.itens_bo):
            tb.Label(frame_itensBo, text=cod).grid(row=idx+1, column=0, padx=5, pady=2)
            tb.Label(frame_itensBo, text=desc).grid(row=idx+1, column=1, padx=5, pady=2)
            tb.Label(frame_itensBo, text=linha).grid(row=idx+1, column=2, padx=5, pady=2)
            entry_motivo = tb.Combobox(frame_itensBo, values=motivos_opcoes, width=25)
            entry_motivo.set(motivo if motivo else "")
            entry_motivo.grid(row=idx+1, column=3, padx=5, pady=2)
            self.entries_motivo_itens.append((cod, entry_motivo))

    def opcoes_editor(self, parent):
        frame = tb.LabelFrame(parent, text="Opções", padding=10)
        frame.pack(side="bottom", fill="x", padx=10, pady=10)

        tb.Button(frame, text="Salvar", command=self.salvar).pack(side="left", padx=5)
        tb.Button(frame, text="Cancelar", command=self.cancel).pack(side="left", padx=5)

    def obter_setores(self):
        conn = create_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT setor_responsavel FROM bo_records WHERE setor_responsavel IS NOT NULL AND setor_responsavel NOT LIKE ''")
        setores = [row[0] for row in cursor.fetchall()]
        cursor.close()
        conn.close()
        return setores

    def obter_motivos(self):
        conn = create_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT MOTIVO FROM BO_ITENS WHERE MOTIVO IS NOT NULL AND MOTIVO NOT LIKE ''")
        motivos = [row[0] for row in cursor.fetchall()]
        cursor.close()
        conn.close()
        return motivos

    def salvar(self):
        try:
            conn = create_connection()

            cursor = conn.cursor()

            update_query = """
                UPDATE bo_records
                SET tipo_ocorrencia = ?,
                    descricao = ?,
                    setor_responsavel = ?,
                    frete = ?
                WHERE bo_number = ?
            """
            
            cursor.execute(update_query, (
                self.entries["tipo_ocorrencia"].get(),
                self.entries["descricao"].get("1.0", tk.END).strip(),
                self.entries["setor_responsavel"].get(),
                self.entries["frete"].get(),
                self.bo_number
            ))

            dados_itens = [
                (entry.get(), self.bo_number, cod)
                for cod, entry in self.entries_motivo_itens
            ]

            cursor.executemany("""
                UPDATE BO_ITENS
                SET MOTIVO = ?
                WHERE BO_REF = ? AND COD = ?
            """, dados_itens)

            conn.commit()
            messagebox.showinfo("Sucesso", f"BO {self.bo_number} atualizada com sucesso!")
            self.root.destroy()
            if exibir_detalhes_acompanhando.instance and exibir_detalhes_acompanhando.instance.root.winfo_exists():
                exibir_detalhes_acompanhando.instance.root.destroy()

            # Atualiza a lista de BOs no módulo que chamou
            if self.ultimo_modulo == "Corporativo":
                from modulos.corporativo import CorporativoModule
                if CorporativoModule.instance is not None:
                    CorporativoModule.instance.atualizar_bos()
            elif self.ultimo_modulo == "Varejo":
                from modulos.varejo import VarejoModule
                if VarejoModule.instance is not None:
                    VarejoModule.instance.atualizar_bos()
            elif self.ultimo_modulo == "Exportacao":
                from modulos.exportacao import ExportacaoModule
                if ExportacaoModule.instance is not None:
                    ExportacaoModule.instance.atualizar_bos()
        except pyodbc.Error as e:
            messagebox.showerror("Erro", f"Erro ao atualizar BO: {e}")
        finally:
            if conn:
                cursor.close()
                conn.close()

    def cancel(self):
        self.root.destroy()

class acompanhar_Bo:
    def __init__(self, parent, tree, caller_id=None, user=None):
        self.user = user
        self.parent = parent
        self.ultimo_modulo = caller_id
        self.tree = tree
        self.anexos = []
        self.loading_label = None
        self.max_anexos = 5
        self.max_tamanho_anexo = 50 * 1024 * 1024

        # ---------- JANELA ----------
        self.root = tk.Toplevel(parent)
        self.root.title("Acompanhar BO")
        self.root.iconbitmap(resource_path('SAREM.ico'))
        self.root.transient(parent)
        self.root.grab_set()
        self.root.focus_set()

        # ---------- SCROLL GLOBAL ----------
        canvas = tk.Canvas(self.root)
        v_scroll = tb.Scrollbar(self.root, orient="vertical", command=canvas.yview)
        v_scroll.pack(side="right", fill="y")
        canvas.pack(side="top", fill="both", expand=True)
        canvas.configure(yscrollcommand=v_scroll.set)

        scrollable_frame = tb.Frame(canvas)
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

        # ---------- CONTEÚDO ROLÁVEL ----------
        self.bo_dados = self.obter_dados_bo()
        self.secao_dados_gerais(scrollable_frame)
        self.secao_ocorrencia(scrollable_frame)
        self.secao_transporte(scrollable_frame)
        self.secao_itens(scrollable_frame)

        # ---------- SCROLL COM A RODA ----------
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # ---------- FRAME DE BOTÕES FIXO ----------
        self.botao_salvar(self.root)   # empacota no root, fora do scroll

        # ---------- AJUSTE DE TAMANHO ----------
        self.root.update_idletasks()
        ideal_width = scrollable_frame.winfo_reqwidth() + v_scroll.winfo_width() + 40
        self.root.geometry(f"{ideal_width}x600")
        center_window_child(self.root, parent)

        # centraliza
        self.root.update_idletasks()
        center_window_child(self.root, parent)

    # ---------- MÉTODOS AUXILIARES ----------
    def obter_dados_bo(self):
        selecionado = self.tree.selection()
        if not selecionado:
            return None
        valores = self.tree.item(selecionado, "values")
        return valores if valores else None

    # ---------- SEÇÕES (todas recebem parent_frame = scrollable_frame) ----------
    def secao_dados_gerais(self, parent_frame):
        frame = tb.LabelFrame(parent_frame, text="Dados Gerais", padding=10)
        frame.pack(fill="x", padx=10, pady=10)

        labels = ["BO:", "OP:", "CLIENTE:", "EMISSÃO:"]
        for i, texto in enumerate(labels):
            tb.Label(frame, text=texto).grid(row=i, column=0, sticky="w")
            tb.Label(frame, text=self.bo_dados[i]).grid(row=i, column=1, sticky="w")

    def obter_setores(self):
        conn = create_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT DISTINCT setor_responsavel FROM bo_records "
            "WHERE setor_responsavel IS NOT NULL AND setor_responsavel NOT LIKE ''"
        )
        setores = [row[0] for row in cursor.fetchall()]
        cursor.close()
        conn.close()
        return setores

    def obter_motivos(self):
        conn = create_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT DISTINCT MOTIVO FROM BO_ITENS "
            "WHERE MOTIVO IS NOT NULL AND MOTIVO NOT LIKE ''"
        )
        motivos = [row[0] for row in cursor.fetchall()]
        cursor.close()
        conn.close()
        return motivos

    def secao_ocorrencia(self, parent_frame):
        frame = tb.LabelFrame(parent_frame, text="Detalhes da Ocorrência", padding=10)
        frame.pack(fill="x", padx=10, pady=10)

        campos = [
            ("Tipo de Ocorrência", tb.Entry),
            ("Setor Responsável", tb.Combobox, self.obter_setores()),
            ("Descrição", tk.Text),
        ]

        self.entries_ocorrencia = {}
        for i, (label, widget, *args) in enumerate(campos):
            tb.Label(frame, text=f"{label}:").grid(row=i, column=0, sticky="w")
            if widget == tk.Text:
                w = widget(frame, width=30, height=5, font=("Arial", 8))
            elif widget == tb.Combobox:
                w = widget(frame, values=args[0])
            else:
                w = widget(frame)
            w.grid(row=i, column=1, sticky="ew")
            self.entries_ocorrencia[label] = w

    def secao_transporte(self, parent_frame):
        frame = tb.LabelFrame(parent_frame, text="Transporte", padding=10)
        frame.pack(fill="x", padx=10, pady=10)

        tb.Label(frame, text="Previsão de entrega:").grid(row=0, column=0, sticky="w")
        tb.Label(frame, text=self.bo_dados[5]).grid(row=0, column=1, sticky="w")

        tb.Label(frame, text="Frete:").grid(row=1, column=0, sticky="w")
        cb_frete = tb.Combobox(frame, values=["CIF", "FOB"], state="readonly")
        cb_frete.grid(row=1, column=1, sticky="w")
        self.entries_transporte = {"Frete": cb_frete}

    def secao_itens(self, parent_frame):
        frame = tb.LabelFrame(parent_frame, text="Itens da BO", padding=10)
        frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Cabeçalho
        headers = ["CÓDIGO", "DESCRIÇÃO", "LINHA", "MOTIVO"]
        for col, header in enumerate(headers):
            tb.Label(frame, text=header, font=("Arial", 9, "bold")).grid(row=0, column=col, padx=5, pady=2)

        # Busca itens no Protheus
        conn = create_connection_Protheus()
        cursor = conn.cursor()
        query = """
            SELECT SC6.C6_CODTIDI, SC6.C6_DESCRI, SC6.C6_LINHA
            FROM SC6010 SC6
            INNER JOIN SC5010 SC5
              ON (SC5.C5_NUM = SC6.C6_NUM
                  AND SC5.C5_FILIAL = SC6.C6_FILIAL
                  AND SC5.D_E_L_E_T_ <> '*')
            WHERE SC6.C6_NUM LIKE ?
              AND SC5.C5_PEDREPR LIKE ?
              AND SC5.C5_PEDREPR LIKE ?
        """
        try:
            cursor.execute(query, (f"%{self.bo_dados[1]}%", "%BO%", f"%{self.bo_dados[0]}%"))
            self.itens_bo = cursor.fetchall()
        finally:
            cursor.close()
            conn.close()

        # Campos de motivo
        self.entries_motivo_itens = []
        motivos = self.obter_motivos()
        for idx, (cod, desc, linha) in enumerate(self.itens_bo, start=1):
            tb.Label(frame, text=cod).grid(row=idx, column=0, padx=5, pady=2)
            tb.Label(frame, text=desc).grid(row=idx, column=1, padx=5, pady=2)
            tb.Label(frame, text=linha).grid(row=idx, column=2, padx=5, pady=2)
            cb = tb.Combobox(frame, values=motivos, width=25)
            cb.grid(row=idx, column=3, padx=5, pady=2)
            self.entries_motivo_itens.append(cb)

    def botao_salvar(self, parent):
        frame = tb.Frame(parent, padding=10)
        frame.pack(side="bottom", fill="x", padx=10, pady=10)
        tb.Button(frame, text="Salvar", command=self.salvar).pack(side="right")

    def salvar(self):
        conn = create_connection()
        if conn is None:
            return
        try:
            cursor = conn.cursor()

            # Verifica se já existe
            cursor.execute(
                "SELECT 1 FROM bo_records WHERE bo_number = ? AND D_E_L_E_T_ <> '*'",
                (self.bo_dados[0],)
            )
            if cursor.fetchone():
                messagebox.showerror("Erro", "BO já cadastrada!")
                self.root.destroy()
                return

            # Insere cabeçalho
            cursor.execute("""
                INSERT INTO bo_records (
                    bo_number, op, loja, filial, emissao_totvs,
                    tipo_ocorrencia, descricao, setor_responsavel,
                    frete, previsao_embarque, modulo, status, user_include
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                self.bo_dados[0], self.bo_dados[1], self.bo_dados[2],
                self.bo_dados[3], self.bo_dados[4],
                self.entries_ocorrencia["Tipo de Ocorrência"].get(),
                self.entries_ocorrencia["Descrição"].get("1.0", tk.END).strip(),
                self.entries_ocorrencia["Setor Responsável"].get(),
                self.entries_transporte["Frete"].get(),
                self.bo_dados[5], self.ultimo_modulo, 'Em Andamento',
                self.user[1]
            ))

            # Itens
            cursor.execute("DELETE FROM BO_ITENS WHERE BO_REF = ?", (self.bo_dados[0],))
            for idx, (cod, desc, linha) in enumerate(self.itens_bo):
                motivo = self.entries_motivo_itens[idx].get()
                cursor.execute("""
                    INSERT INTO BO_ITENS (BO_REF, COD, [DESC], LINHA, MOTIVO)
                    VALUES (?, ?, ?, ?, ?)
                """, (self.bo_dados[0], cod, desc, linha, motivo))

            conn.commit()
            messagebox.showinfo("Sucesso", "BO salva com sucesso!")
            self.root.destroy()
            if exibir_detalhes.instance and exibir_detalhes.instance.root.winfo_exists():
                exibir_detalhes.instance.root.destroy()

            # Atualiza lista
            if self.ultimo_modulo == "Corporativo":
                from modulos.corporativo import CorporativoModule
                if CorporativoModule.instance:
                    CorporativoModule.instance.atualizar_bos()
            elif self.ultimo_modulo == "Varejo":
                from modulos.varejo import VarejoModule
                if VarejoModule.instance:
                    VarejoModule.instance.atualizar_bos()
            elif self.ultimo_modulo == "Exportacao":
                from modulos.exportacao import ExportacaoModule
                if ExportacaoModule.instance:
                    ExportacaoModule.instance.atualizar_bos()

        except pyodbc.Error as e:
            messagebox.showerror("Erro", f"Erro ao salvar BO: {e}")
            print(e)
        finally:
            if conn:
                cursor.close()
                conn.close()

    def identificar_chamador(self):
        if self.ultimo_modulo is None:
            raise ValueError("É necessário informar o identificador do módulo que chamou a função")
        return f"Tela de Acompanhar BO chamada pelo módulo: {self.ultimo_modulo}"