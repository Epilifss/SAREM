from back.database import create_connection
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph, Spacer, Image
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from back.utils import resource_path
import pandas as pd
import tempfile
import sys
import os
import time
import subprocess

def generate_report(modulo, cliente, setor, data_inicio, data_fim, status):
    conn = create_connection()
    if not conn:
        return []

    query = """
        SELECT
            BR.bo_number,
            BR.op,
            BR.loja,
            BR.tipo_ocorrencia,
            BR.setor_responsavel,
            BR.[status],
            BR.created_at,
            ISNULL(ITENS.produtos_motivos, '') AS produtos_motivos
        FROM bo_records AS BR
        OUTER APPLY (
            SELECT STUFF((
                SELECT CHAR(10)
                    + COALESCE(BI.COD, '')
                    + ' - '
                    + COALESCE(BI.[DESC], '')
                    + ' | Motivo: '
                    + COALESCE(BI.MOTIVO, '')
                FROM BO_ITENS AS BI
                WHERE BI.BO_REF = BR.bo_number
                FOR XML PATH(''), TYPE
            ).value('.', 'NVARCHAR(MAX)'), 1, 1, '') AS produtos_motivos
        ) AS ITENS
        WHERE BR.modulo LIKE ?
            AND BR.D_E_L_E_T_ <> '*'
    """
    params = [modulo]

    if cliente != "Todos":
        query += " AND BR.loja = ?"
        params.append(cliente)
    if setor != "Todos":
        query += " AND BR.setor_responsavel = ?"
        params.append(setor)
    query += " AND CAST(BR.emissao_totvs AS DATE) BETWEEN ? AND ?"
    params.extend([data_inicio, data_fim])
    if status != "Todos":
        query += " AND BR.[status] = ?"
        params.append(status)

    query += " ORDER BY BR.created_at DESC"

    try:
        with conn.cursor() as cursor:
            cursor.execute(query, params)
            return cursor.fetchall()
    finally:
        conn.close()

def get_all_clients(ultimo_modulo):
    conn = create_connection()
    if not conn:
        return ["Todos"]

    query = "SELECT COALESCE(loja, 'Não especificado') AS Setor FROM bo_records WHERE loja IS NOT NULL AND loja NOT LIKE '' AND modulo LIKE ? AND D_E_L_E_T_ <> '*' GROUP BY loja"
    try:
        with conn.cursor() as cursor:
            cursor.execute(query, ultimo_modulo)
            return ["Todos"] + [row[0] for row in cursor.fetchall()]
    finally:
        conn.close()

def get_all_setores(ultimo_modulo):
    conn = create_connection()
    if not conn:
        return ["Todos"]

    query = "SELECT COALESCE(setor_responsavel, 'Não especificado') AS Setor FROM bo_records WHERE setor_responsavel IS NOT NULL AND setor_responsavel NOT LIKE '' AND modulo LIKE ? AND D_E_L_E_T_ <> '*' GROUP BY setor_responsavel"
    try:
        with conn.cursor() as cursor:
            cursor.execute(query, ultimo_modulo)
            return ["Todos"] + [row[0] for row in cursor.fetchall()]
    finally:
        conn.close()

def get_all_status(ultimo_modulo):
    conn = create_connection()
    if not conn:
        return ["Todos"]

    query = "SELECT COALESCE([status], 'Não especificado') AS [Status] FROM bo_records WHERE [status] IS NOT NULL AND [status] NOT LIKE '' AND modulo LIKE ? AND D_E_L_E_T_ <> '*' GROUP BY [status]"
    try:
        with conn.cursor() as cursor:
            cursor.execute(query, ultimo_modulo)
            return ["Todos"] + [row[0] for row in cursor.fetchall()]
    finally:
        conn.close()

def export_to_pdf(data, headers, file_path):
    doc = SimpleDocTemplate(file_path, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
        "HeaderCell",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=7,
        leading=8,
        textColor=colors.white,
        alignment=1,
    )
    body_style = ParagraphStyle(
        "BodyCell",
        parent=styles["BodyText"],
        fontSize=8,
        leading=10,
    )

    logo_path = resource_path('SAREM PNG.png')
    if os.path.exists(logo_path):
        logo = Image(logo_path, width=120, height=38)
        logo.hAlign = 'LEFT'
        elements.append(logo)

    elements.append(Paragraph("Relatório de BOs", styles['Title']))
    elements.append(Paragraph(time.strftime("Gerado em: %d/%m/%Y %H:%M"), styles['Normal']))
    elements.append(Spacer(1, 12))

    header_alias = {
        "TIPO DE OCORRÊNCIA": "TIPO DE<br/>OCORRÊNCIA",
        "SETOR RESPONSÁVEL": "SETOR<br/>RESPONSÁVEL",
        "DATA DE REGISTRO": "DATA DE<br/>REGISTRO",
        "PRODUTOS / MOTIVOS": "PRODUTOS /<br/>MOTIVOS",
    }
    formatted_headers = [
        Paragraph(header_alias.get(h, h), header_style)
        for h in headers
    ]

    table_data = [formatted_headers]
    for row in data:
        formatted_row = []
        for value in row:
            text = "" if value is None else str(value).replace("\n", "<br/>")
            formatted_row.append(Paragraph(text, body_style))
        table_data.append(formatted_row)

    col_widths = [45, 45, 75, 85, 95, 55, 70, 230]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#24577f')),
        ('ALIGN', (0, 0), (6, -1), 'CENTER'),
        ('ALIGN', (7, 0), (7, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 4),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    elements.append(table)

    doc.build(elements)
    return True

def export_to_excel(data, headers, file_path):
    df = pd.DataFrame(data, columns=headers)
    start_row = 4
    df.to_excel(file_path, index=False, startrow=start_row)

    # Formatação avançada
    wb = load_workbook(file_path)
    if wb.sheetnames:
        ws = wb[wb.sheetnames[0]]
    else:
        raise ValueError("No worksheets found in the Excel file.")

    logo_path = resource_path('SAREM PNG.png')
    if os.path.exists(logo_path):
        logo = XLImage(logo_path)
        logo.width = 120
        logo.height = 38
        ws.add_image(logo, 'A1')

    ws['C1'] = 'Relatório de BOs'
    ws['C2'] = time.strftime('Gerado em: %d/%m/%Y %H:%M')
    ws['C1'].font = Font(bold=True, size=14)
    ws['C2'].font = Font(size=10)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    row_fills = [PatternFill(start_color="FFFFFF", fill_type="solid"),
                 PatternFill(start_color="FFFFE0", fill_type="solid")]
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    header_row = start_row + 1
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row, max_row=ws.max_row), start=header_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            if row_idx == header_row:
                cell.font = header_font
                cell.fill = header_fill
            else:
                cell.fill = row_fills[(row_idx - header_row - 1) % 2]

    from openpyxl.utils import get_column_letter

    for idx, col in enumerate(ws.iter_cols(min_row=header_row, max_row=ws.max_row), 1):
        max_length = 0
        column = get_column_letter(idx)
        for cell in col:
            if cell.value:
                lines = str(cell.value).split('\n')
                current_length = max(len(line) for line in lines)
                if current_length > max_length:
                    max_length = current_length
        ws.column_dimensions[column].width = min(max(max_length + 2, 12), 80)

    ws.row_dimensions[1].height = 34
    ws.freeze_panes = f"A{header_row + 1}"

    wb.save(file_path)
    return True

def print_report(data, headers):
    temp_dir = tempfile.gettempdir()
    temp_pdf_path = os.path.join(temp_dir, "temp_report.pdf")

    export_to_pdf(data, headers, temp_pdf_path)

    if sys.platform == "win32":
        os.startfile(temp_pdf_path, "print")
    else:
        subprocess.run(["lp", temp_pdf_path])

    time.sleep(7)

    if os.path.exists(temp_pdf_path):
        os.unlink(temp_pdf_path)
    return True